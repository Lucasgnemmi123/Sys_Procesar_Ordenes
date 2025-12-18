import os
import re
import fitz  # PyMuPDF
import pandas as pd
import xlwings as xw
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- Funciones auxiliares ---
def clean_qty(qtext):
    q = str(qtext).strip()
    if '.' in q and ',' in q:
        q = q.replace('.', '').replace(',', '.')
    else:
        q = q.replace(',', '.')
    try:
        return int(float(q))
    except:
        return q

def extract_centro_costo_y_nombre(text):
    m = re.search(r'Nombre\s*:\s*(\d{3,6})\s+(.*)', text, flags=re.IGNORECASE)
    if m:
        return m.group(1), m.group(2).strip()
    return "", ""

def extract_items_from_text(text):
    pattern = re.compile(
        r'^\s*\d+\s+\d+\s+\d{2}/\d{2}/\d{2}\s+(A\d{4,6})\s+(?:A\d{4,6}\s+)?([\d\.\,]+)',
        re.MULTILINE
    )
    return [{"SKU": sku, "CANTIDAD_RAW": qty} for sku, qty in pattern.findall(text)]

# --- Procesamiento de PDFs ---
def procesar_pdfs(ordenes_dir):
    if not os.path.exists(ordenes_dir):
        raise FileNotFoundError(f"No se encontró la carpeta: {ordenes_dir}")

    rows = []
    for fname in sorted(os.listdir(ordenes_dir)):
        if not fname.lower().endswith('.pdf'):
            continue
        path = os.path.join(ordenes_dir, fname)
        with fitz.open(path) as doc:
            text = "".join([p.get_text("text") + "\n" for p in doc])

        centro_costo, nombre_lugar = extract_centro_costo_y_nombre(text)

        for it in extract_items_from_text(text):
            qty = clean_qty(it["CANTIDAD_RAW"])
            rows.append({
                "LOCAL": "00000",
                "SKU": str(it["SKU"]).strip().upper(),
                "CANTIDAD": qty,
                "CENTRO_COSTO": centro_costo,
                "NOMBRE_LUGAR": nombre_lugar,
                "_SRC_FILE": fname
            })

    return pd.DataFrame(rows).reset_index(drop=True)

# --- Validar SKUs contra Items.xlsx ---
def validar_skus_items(df, items_xlsx):
    """Valida que los SKUs estén en la lista de items de compra calzada"""
    df = df.copy()
    df_err = pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
    warnings = []

    try:
        # Leer archivo Items.xlsx
        df_items = pd.read_excel(items_xlsx, dtype=str)
        df_items = df_items.rename(columns=lambda x: str(x).strip())
        
        # Buscar columna SKU en Items
        col_sku_items = next((c for c in df_items.columns if "sku" in c.lower() or "codigo" in c.lower()), None)
        
        if not col_sku_items:
            warnings.append(f"⚠️ Columna SKU no encontrada en Items.xlsx, saltando validación")
            return df, df_err, warnings
        
        # Preparar lista de SKUs válidos
        skus_validos = set(df_items[col_sku_items].astype(str).str.strip().str.upper())
        warnings.append(f"✅ Items cargados: {len(skus_validos)} SKUs válidos encontrados")
        
        # Validar cada SKU
        mask_no_en_items = ~df["SKU"].isin(skus_validos)
        
        if mask_no_en_items.any():
            df_err_items = df.loc[mask_no_en_items].copy()
            df_err_items["OBSERVACION"] = df_err_items["CENTRO_COSTO"].fillna("") + "//No está en items de compra calzada//" + df_err_items["NOMBRE_LUGAR"].fillna("")
            df_err = pd.concat([df_err, df_err_items], ignore_index=True)
            warnings.append(f"⚠️ {len(df_err_items)} SKUs no encontrados en items de compra calzada")
        
        # Retornar solo los válidos
        df_valid = df.loc[~mask_no_en_items].copy()
        
    except Exception as e:
        warnings.append(f"❌ Error leyendo Items.xlsx: {e}")
        df_valid = df.copy()  # Si hay error, continuar sin validación
        
    return df_valid.reset_index(drop=True), df_err.reset_index(drop=True), warnings

# --- Mapear proveedor desde Full.xlsx ---
def mapear_proveedor_por_sku(df, full_xlsx, region="099"):
    df = df.copy()
    df_err = pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
    warnings = []

    try:
        df_full = pd.read_excel(full_xlsx, sheet_name='Full', dtype=str)
        df_full = df_full.rename(columns=lambda x: str(x).strip())
        col_sku = next((c for c in df_full.columns if "codigo sku" in c.lower() or "sku" in c.lower()), None)
        col_prov = next((c for c in df_full.columns if "código proveedor" in c.lower() or "proveedor" in c.lower()), None)
        col_region = next((c for c in df_full.columns if "región" in c.lower() or "region" in c.lower()), None)

        if not col_sku or not col_prov:
            warnings.append(f"❌ Columnas SKU o PROVEEDOR no encontradas en Full.xlsx")
            df["PROVEEDOR"] = None
            return df, df_err, warnings

        # Seleccionar columnas necesarias
        columnas_necesarias = [col_sku, col_prov]
        if col_region:
            columnas_necesarias.append(col_region)
            df_full = df_full[columnas_necesarias].rename(columns={col_sku: "SKU", col_prov: "PROVEEDOR", col_region: "REGION"})
            
            # Filtrar por región si la columna existe
            df_full["REGION"] = df_full["REGION"].astype(str).str.strip()
            df_full = df_full[df_full["REGION"] == region]
            warnings.append(f"✅ Filtrado por región: {region} ({len(df_full)} registros encontrados)")
        else:
            df_full = df_full[columnas_necesarias].rename(columns={col_sku: "SKU", col_prov: "PROVEEDOR"})
            warnings.append(f"⚠️ Columna REGIÓN no encontrada, procesando sin filtro de región")

        df_full["SKU"] = df_full["SKU"].astype(str).str.strip().str.upper()
        df_full["PROVEEDOR"] = df_full["PROVEEDOR"].astype(str).str.strip()
        
        # Eliminar duplicados en la tabla de referencia - tomar el primer proveedor para cada SKU
        df_full = df_full.drop_duplicates(subset=['SKU'], keep='first')

        # merge
        df = df.merge(df_full[["SKU", "PROVEEDOR"]], on="SKU", how="left")

        # marcar errores "Falta precio"
        mask_falta_precio = df["PROVEEDOR"].isna() | (df["PROVEEDOR"] == "") | (df["PROVEEDOR"] == "nan")
        if mask_falta_precio.any():
            df_err_precio = df.loc[mask_falta_precio].copy()
            df_err_precio["OBSERVACION"] = df_err_precio["CENTRO_COSTO"].fillna("") + "//Falta precio//" + df_err_precio["NOMBRE_LUGAR"].fillna("")
            df_err = pd.concat([df_err, df_err_precio], ignore_index=True)

        df_valid = df.loc[~mask_falta_precio].copy()

    except Exception as e:
        warnings.append(f"❌ Error leyendo Full.xlsx: {e}")
        df["PROVEEDOR"] = None
        df_valid = df.copy()

    return df_valid.reset_index(drop=True), df_err.reset_index(drop=True), warnings

# --- Rellenar fecha, observación y asignar ID ---
def rellenar_fecha_entrega_y_observacion(df, agenda_xlsm):
    df = df.copy()
    df_err = pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
    try:
        app = xw.App(visible=False)
        wb = app.books.open(agenda_xlsm)
        ws_matriz = wb.sheets["Matriz"]
        ultima_fila = ws_matriz.range("A" + str(ws_matriz.cells.last_cell.row)).end("up").row
        data_matriz = ws_matriz.range(f"A3:K{ultima_fila}").value if ultima_fila >= 3 else []
        columnas = ["PROVEEDOR","NOMBRE","VACIA","LUN","MAR","MIE","JUE","VIE","SAB","D-1","ENTREGA"]
        df_matriz = pd.DataFrame(data_matriz, columns=columnas) if data_matriz else pd.DataFrame(columns=columnas)

        df_matriz["PROVEEDOR"] = df_matriz["PROVEEDOR"].astype(str).str.strip()
        df_matriz["ENTREGA"] = pd.to_datetime(df_matriz["ENTREGA"], format="%d-%m-%Y", errors="coerce")
        df_matriz["FECHA_ENTREGA"] = df_matriz["ENTREGA"].dt.strftime("%d-%m-%Y")

        df = df.merge(df_matriz[["PROVEEDOR","FECHA_ENTREGA"]], on="PROVEEDOR", how="left")
        mask_falta_agenda = df["FECHA_ENTREGA"].isna()
        if mask_falta_agenda.any():
            df_err_agenda = df.loc[mask_falta_agenda].copy()
            df_err_agenda["OBSERVACION"] = df_err_agenda["CENTRO_COSTO"].fillna("") + "//Falta Agenda//" + df_err_agenda["NOMBRE_LUGAR"].fillna("")
            df_err = pd.concat([df_err, df_err_agenda], ignore_index=True)

        df_valid = df.loc[~mask_falta_agenda].copy()
        fecha_despacho = ws_matriz.range("M2").value
        if not isinstance(fecha_despacho, datetime):
            try: fecha_despacho = pd.to_datetime(fecha_despacho)
            except: fecha_despacho = datetime.today()
        dd_mm = fecha_despacho.strftime("%d-%m")
        df_valid["OBSERVACION"] = df_valid["CENTRO_COSTO"].fillna("") + f"//{dd_mm}//BOD. " + df_valid["NOMBRE_LUGAR"].fillna("")

    finally:
        wb.close()
        app.quit()

    return df_valid.reset_index(drop=True), df_err.reset_index(drop=True)

# --- Obtener fecha de M1 y generar nombre de archivo ---
def obtener_nombre_archivo_salida(agenda_xlsm, base_dir):
    """Genera el nombre del archivo de salida basado en la fecha de M1 en Agenda.xlsm"""
    try:
        app = xw.App(visible=False)
        wb = app.books.open(agenda_xlsm)
        ws_matriz = wb.sheets["Matriz"]
        
        # Leer fecha de M1
        fecha_m1 = ws_matriz.range("M1").value
        if not isinstance(fecha_m1, datetime):
            try: 
                fecha_m1 = pd.to_datetime(fecha_m1)
            except: 
                fecha_m1 = datetime.today()
        
        # Formatear fecha como DD-MM-YYYY
        fecha_str = fecha_m1.strftime("%d-%m-%Y")
        nombre_archivo = f"PEDIDOS_CD_OVIEDO_{fecha_str}.xlsx"
        
        # Crear carpeta Salidas si no existe
        salidas_dir = os.path.join(base_dir, "Salidas")
        os.makedirs(salidas_dir, exist_ok=True)
        
        return os.path.join(salidas_dir, nombre_archivo)
        
    except Exception as e:
        print(f"⚠️ Error leyendo fecha de M1, usando fecha actual: {e}")
        fecha_actual = datetime.today().strftime("%d-%m-%Y")
        nombre_archivo = f"PEDIDOS_CD_OVIEDO_{fecha_actual}.xlsx"
        
        # Crear carpeta Salidas si no existe
        salidas_dir = os.path.join(base_dir, "Salidas")
        os.makedirs(salidas_dir, exist_ok=True)
        
        return os.path.join(salidas_dir, nombre_archivo)
        
    finally:
        try:
            wb.close()
            app.quit()
        except:
            pass

# --- Asignar IDs finales ---
def asignar_id_final(df):
    df = df.copy()
    
    # Primero eliminar duplicados y consolidar cantidades
    print("🔄 Eliminando duplicados y consolidando cantidades...")
    
    # Identificar columnas clave para agrupación (sin incluir cantidad)
    columnas_agrupacion = ["LOCAL", "SKU", "PROVEEDOR", "FECHA_ENTREGA", "OBSERVACION"]
    
    # Verificar que todas las columnas existan
    columnas_existentes = [col for col in columnas_agrupacion if col in df.columns]
    
    if len(columnas_existentes) == len(columnas_agrupacion):
        # Agrupar por las columnas clave y sumar las cantidades
        df_consolidado = df.groupby(columnas_agrupacion, as_index=False).agg({
            'CANTIDAD': 'sum',  # Sumar las cantidades
            'CENTRO_COSTO': 'first',  # Tomar el primer valor (debería ser igual)
            'NOMBRE_LUGAR': 'first',  # Tomar el primer valor (debería ser igual)
            '_SRC_FILE': 'first'  # Tomar el primer archivo (para referencia)
        })
        
        print(f"✅ Consolidados {len(df)} registros a {len(df_consolidado)} registros únicos")
        df = df_consolidado
    else:
        print(f"⚠️ No se pudo consolidar: faltan columnas {set(columnas_agrupacion) - set(columnas_existentes)}")
    
    # Ahora asignar IDs
    df = df.sort_values(["PROVEEDOR","OBSERVACION","SKU"]).reset_index(drop=True)
    ids, next_id, last_pair = [], 1, (None, None)
    for _, row in df.iterrows():
        pair = (row["PROVEEDOR"], row["OBSERVACION"])
        if pair != last_pair:
            ids.append(next_id)
            last_pair = pair
            next_id += 1
        else:
            ids.append(next_id-1)
    df["ID PEDIDO"] = ids
    
    # Reordenar columnas según el formato requerido
    columnas_finales = ["ID PEDIDO", "LOCAL", "PROVEEDOR", "FECHA_ENTREGA", "SKU", "CANTIDAD", "OBSERVACION"]
    
    # Asegurar que todas las columnas existan
    for col in columnas_finales:
        if col not in df.columns:
            df[col] = ""
    
    # Seleccionar solo las columnas requeridas en el orden correcto
    df = df[columnas_finales]
    
    return df

# --- Formatear archivo Excel de salida ---
def formatear_excel_salida(archivo_excel):
    """Formatear el archivo Excel con anchos de columna y estilos"""
    try:
        # Cargar el archivo
        wb = load_workbook(archivo_excel)
        
        # Formatear hoja "PEDIDOS_CD"
        if "PEDIDOS_CD" in wb.sheetnames:
            ws_salida = wb["PEDIDOS_CD"]
            
            # Definir anchos de columna optimizados
            anchos_columnas = {
                'A': 12,  # ID PEDIDO
                'B': 8,   # LOCAL
                'C': 12,  # PROVEEDOR
                'D': 15,  # FECHA_ENTREGA
                'E': 10,  # SKU
                'F': 10,  # CANTIDAD
                'G': 60   # OBSERVACION (más ancho para texto largo)
            }
            
            # Aplicar anchos de columna
            for col, ancho in anchos_columnas.items():
                ws_salida.column_dimensions[col].width = ancho
            
            # Formatear encabezados
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="2F5597", end_color="2F5597", fill_type="solid")
            header_alignment = Alignment(horizontal="center", vertical="center")
            
            # Aplicar formato a la primera fila (encabezados)
            for cell in ws_salida[1]:
                if cell.value:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment
            
            # Formatear datos
            data_alignment_center = Alignment(horizontal="center", vertical="center")
            data_alignment_left = Alignment(horizontal="left", vertical="center")
            
            # Aplicar alineación a las columnas de datos
            for row in ws_salida.iter_rows(min_row=2):
                for i, cell in enumerate(row):
                    if i in [0, 1, 2, 3, 4, 5]:  # ID, LOCAL, PROVEEDOR, FECHA, SKU, CANTIDAD
                        cell.alignment = data_alignment_center
                    elif i == 6:  # OBSERVACION
                        cell.alignment = data_alignment_left
            
            print("✅ Formato aplicado a hoja 'PEDIDOS_CD'")
        
        # Formatear hoja "Errores" si existe
        if "Errores" in wb.sheetnames:
            ws_errores = wb["Errores"]
            
            # Anchos para hoja de errores
            anchos_errores = {
                'A': 8,   # LOCAL
                'B': 10,  # SKU
                'C': 10,  # CANTIDAD
                'D': 12,  # CENTRO_COSTO
                'E': 30,  # NOMBRE_LUGAR
                'F': 20,  # _SRC_FILE
                'G': 50   # OBSERVACION
            }
            
            # Aplicar anchos
            for col, ancho in anchos_errores.items():
                ws_errores.column_dimensions[col].width = ancho
            
            # Formatear encabezados de errores
            header_font_error = Font(bold=True, color="FFFFFF")
            header_fill_error = PatternFill(start_color="C5504B", end_color="C5504B", fill_type="solid")
            
            for cell in ws_errores[1]:
                if cell.value:
                    cell.font = header_font_error
                    cell.fill = header_fill_error
                    cell.alignment = header_alignment
            
            # Alineación para datos de errores
            for row in ws_errores.iter_rows(min_row=2):
                for cell in row:
                    cell.alignment = data_alignment_left
            
            print("✅ Formato aplicado a hoja 'Errores'")
        
        # Guardar cambios
        wb.save(archivo_excel)
        print(f"✅ Archivo formateado guardado: {archivo_excel}")
        
    except Exception as e:
        print(f"⚠️ Error al formatear Excel: {e}")
        print("📄 El archivo se guardó sin formato especial")
