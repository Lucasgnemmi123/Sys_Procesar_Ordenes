"""
Sistema de Gestión de Agenda - Reemplazo de Agenda.xlsm
Creado por Lucas Gnemmi
Versión: 1.0

Gestiona la matriz de proveedores y calcula fechas de entrega automáticamente.
Reemplaza la funcionalidad del Excel Agenda.xlsm con lógica Python pura.
"""

import json
import os
from datetime import datetime, timedelta
from typing import Dict, List, Tuple, Optional


class AgendaManager:
    """Gestor de agenda de proveedores y cálculo de fechas de entrega"""
    
    def __init__(self, config_file: str = None):
        """
        Inicializa el gestor de agenda
        
        Args:
            config_file: Ruta al archivo de configuración JSON (opcional)
        """
        # Si no se especifica, usar ubicación por defecto
        if config_file is None:
            # Guardar en la carpeta raíz del proyecto
            base_dir = os.path.dirname(os.path.abspath(__file__))
            self.config_file = os.path.join(base_dir, "agenda_config.json")
        else:
            self.config_file = config_file
            
        self.dias_despacho = 21  # Días por defecto para calcular fecha de despacho
        self.proveedores = {}  # Diccionario de proveedores y sus días de entrega
        self.cargar_configuracion()
    
    def cargar_configuracion(self):
        """Carga la configuración desde el archivo JSON"""
        if os.path.exists(self.config_file):
            try:
                with open(self.config_file, 'r', encoding='utf-8') as f:
                    data = json.load(f)
                    self.dias_despacho = data.get('dias_despacho', 21)
                    self.proveedores = data.get('proveedores', {})
                    # Configuración cargada exitosamente
            except Exception as e:
                # Error cargando configuración
                self._crear_configuracion_default()
        else:
            # Creando configuración por defecto
            self._crear_configuracion_default()
    
    def _crear_configuracion_default(self):
        """Crea configuración por defecto"""
        self.dias_despacho = 21
        self.proveedores = {}
        self.guardar_configuracion()
    
    def guardar_configuracion(self):
        """Guarda la configuración en el archivo JSON"""
        try:
            data = {
                'dias_despacho': self.dias_despacho,
                'proveedores': self.proveedores
            }
            with open(self.config_file, 'w', encoding='utf-8') as f:
                json.dump(data, f, indent=4, ensure_ascii=False)
            # Configuración guardada
        except Exception as e:
            pass  # Error guardando configuración
    
    def agregar_proveedor(self, codigo: str, nombre: str, dias_entrega: Dict[str, any], dias_d2: any = None, fecha_manual: str = None):
        """
        Agrega o actualiza un proveedor en la matriz
        
        Args:
            codigo: Código del proveedor
            nombre: Nombre del proveedor
            dias_entrega: Diccionario con días de la semana {LUN: 1/0/None, MAR: 1/0/None, ...}
                         1 = Aplica para entrega
                         0 = No aplica esta vez (pero podría aplicar, es recordatorio)
                         None = Ignorar (no se usa este día)
            dias_d2: Si entrega D-2 (1/0/None)
            fecha_manual: Fecha de entrega manual en formato dd-mm-yyyy (opcional)
        """
        self.proveedores[str(codigo)] = {
            'nombre': nombre,
            'LUN': dias_entrega.get('LUN', None),
            'MAR': dias_entrega.get('MAR', None),
            'MIE': dias_entrega.get('MIE', None),
            'JUE': dias_entrega.get('JUE', None),
            'VIE': dias_entrega.get('VIE', None),
            'SAB': dias_entrega.get('SAB', None),
            'D-2': dias_d2,
            'fecha_manual': fecha_manual
        }
        self.guardar_configuracion()
    
    def eliminar_proveedor(self, codigo: str):
        """Elimina un proveedor de la matriz"""
        if str(codigo) in self.proveedores:
            del self.proveedores[str(codigo)]
            self.guardar_configuracion()
            return True
        return False
    
    def obtener_proveedor(self, codigo: str) -> Optional[Dict]:
        """Obtiene la configuración de un proveedor"""
        return self.proveedores.get(str(codigo))
    
    def obtener_todos_proveedores(self) -> Dict:
        """Retorna todos los proveedores"""
        return self.proveedores.copy()
    
    def calcular_fecha_despacho(self, fecha_pedido: datetime) -> datetime:
        """
        Calcula la fecha de despacho sumando los días configurados
        
        Args:
            fecha_pedido: Fecha del pedido
            
        Returns:
            Fecha de despacho calculada
        """
        return fecha_pedido + timedelta(days=self.dias_despacho)
    
    def calcular_fecha_entrega(self, codigo_proveedor: str, fecha_despacho: datetime) -> Optional[datetime]:
        """
        Calcula la fecha de entrega para un proveedor específico
        Implementa la lógica del VBA de Agenda.xlsm
        
        Args:
            codigo_proveedor: Código del proveedor
            fecha_despacho: Fecha de despacho
            
        Returns:
            Fecha de entrega calculada o None si no hay configuración
        """
        proveedor = self.obtener_proveedor(codigo_proveedor)
        if not proveedor:
            return None
        
        # Si tiene fecha manual, usarla
        fecha_manual = proveedor.get('fecha_manual')
        if fecha_manual:
            try:
                return datetime.strptime(fecha_manual, "%d-%m-%Y")
            except:
                pass  # Si hay error, continuar con cálculo automático
        
        # Día de la semana del despacho (0=Lun, 1=Mar, ..., 6=Dom)
        dia_semana_despacho = fecha_despacho.weekday()  # 0=Monday, 6=Sunday
        
        fecha_seleccionada = None
        
        # Mapeo de días de la semana
        dias = ['LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB']
        
        # Revisar D-2 primero (solo si es 1)
        d2_value = proveedor.get('D-2')
        if d2_value == 1 or d2_value is True:
            fecha_test = fecha_despacho - timedelta(days=2)
            if fecha_test < fecha_despacho:
                fecha_seleccionada = fecha_test
        
        # Revisar días de la semana (LUN-SAB) - solo considerar valor 1
        for k, dia in enumerate(dias):
            dia_value = proveedor.get(dia)
            if dia_value == 1 or dia_value is True:  # Solo considerar si es 1 o True
                # Calcular fecha de la semana anterior para ese día
                dias_atras = 7 - (k - dia_semana_despacho)
                if dias_atras <= 0:
                    dias_atras += 7
                fecha_test = fecha_despacho - timedelta(days=dias_atras)
                
                # Solo consideramos fechas anteriores al despacho
                if fecha_test < fecha_despacho:
                    # Lógica según día de despacho
                    if dia_semana_despacho in [0, 1, 2]:  # Lunes, Martes, Miércoles
                        # Tomar la más lejana (mínima)
                        if fecha_seleccionada is None or fecha_test < fecha_seleccionada:
                            fecha_seleccionada = fecha_test
                    elif dia_semana_despacho in [3, 4, 5]:  # Jueves, Viernes, Sábado
                        # Tomar la más próxima (máxima)
                        if fecha_seleccionada is None or fecha_test > fecha_seleccionada:
                            fecha_seleccionada = fecha_test
        
        return fecha_seleccionada
    
    def procesar_dataframe_con_fechas(self, df, fecha_pedido: datetime = None):
        """
        Procesa un DataFrame agregando fechas de entrega para cada proveedor
        
        Args:
            df: DataFrame con columna 'PROVEEDOR'
            fecha_pedido: Fecha del pedido (usa fecha actual si no se proporciona)
            
        Returns:
            DataFrame con columnas FECHA_ENTREGA y OBSERVACION agregadas
        """
        import pandas as pd
        
        if fecha_pedido is None:
            fecha_pedido = datetime.now()
        
        # Calcular fecha de despacho
        fecha_despacho = self.calcular_fecha_despacho(fecha_pedido)
        dd_mm = fecha_despacho.strftime("%d-%m")
        
        df = df.copy()
        df['FECHA_ENTREGA'] = None
        
        # Procesar cada fila
        for idx, row in df.iterrows():
            codigo_prov = str(row.get('PROVEEDOR', '')).strip()
            
            # Normalizar código (eliminar .0 si existe)
            codigo_prov = codigo_prov.replace('.0', '')
            
            if codigo_prov:
                fecha_entrega = self.calcular_fecha_entrega(codigo_prov, fecha_despacho)
                if fecha_entrega:
                    df.at[idx, 'FECHA_ENTREGA'] = fecha_entrega
                else:
                    # Proveedor no configurado
                    pass  # print(f"Proveedor {codigo_prov} no configurado en agenda")
        
        # Generar observaciones
        if 'OBSERVACION' not in df.columns:
            df['OBSERVACION'] = ''
        
        # Agregar fecha de despacho a observaciones
        centro_costo = df.get('CENTRO_COSTO', '').fillna('')
        nombre_lugar = df.get('NOMBRE_LUGAR', '').fillna('')
        
        df['OBSERVACION'] = centro_costo + "//" + dd_mm + "//" + nombre_lugar
        
        return df
    
    def importar_desde_excel(self, ruta_agenda_xlsm: str):
        """
        Importa la configuración desde un archivo Agenda.xlsm existente
        
        Args:
            ruta_agenda_xlsm: Ruta al archivo Agenda.xlsm
        """
        try:
            from openpyxl import load_workbook
            
            # print(f"Importando configuración desde {ruta_agenda_xlsm}...")
            wb = load_workbook(ruta_agenda_xlsm, read_only=True, data_only=True)
            
            if 'Matriz' not in wb.sheetnames:
                # print("No se encontró la hoja 'Matriz'")
                return False
            
            ws = wb['Matriz']
            
            # Leer días de despacho si está configurado (asumimos que puede estar en alguna celda)
            # Por ahora mantenemos el valor actual
            
            # Leer matriz de proveedores (A3:K en adelante)
            proveedores_importados = 0
            for row in ws.iter_rows(min_row=3, values_only=True):
                if row[0]:  # Si hay código de proveedor
                    codigo = str(row[0]).strip().replace('.0', '')
                    nombre = str(row[1]) if row[1] else ''
                    
                    # row[2] es columna vacía
                    # Convertir valores a 1, 0 o None
                    def convertir_valor(val):
                        if val is None or val == '':
                            return None
                        if val == 0:
                            return 0
                        if val == 1 or val:
                            return 1
                        return None
                    
                    dias_entrega = {
                        'LUN': convertir_valor(row[3]),
                        'MAR': convertir_valor(row[4]),
                        'MIE': convertir_valor(row[5]),
                        'JUE': convertir_valor(row[6]),
                        'VIE': convertir_valor(row[7]),
                        'SAB': convertir_valor(row[8]),
                    }
                    dias_d2 = convertir_valor(row[9])
                    
                    self.agregar_proveedor(codigo, nombre, dias_entrega, dias_d2)
                    proveedores_importados += 1
            
            wb.close()
            # print(f"Importados {proveedores_importados} proveedores desde Excel")
            return True
            
        except Exception as e:
            # print(f"Error importando desde Excel: {e}")
            return False


# Función de utilidad para uso rápido
def obtener_gestor_agenda() -> AgendaManager:
    """Retorna una instancia del gestor de agenda"""
    return AgendaManager()


if __name__ == "__main__":
    # Prueba del módulo
    # print("Probando AgendaManager...")
    
    manager = AgendaManager()
    
    # Agregar proveedor de ejemplo
    manager.agregar_proveedor(
        codigo="12345",
        nombre="Proveedor de Prueba",
        dias_entrega={'LUN': True, 'MIE': True, 'VIE': True},
        dias_d2=False
    )
    
    # Calcular fechas
    fecha_pedido = datetime(2026, 1, 6)
    fecha_despacho = manager.calcular_fecha_despacho(fecha_pedido)
    fecha_entrega = manager.calcular_fecha_entrega("12345", fecha_despacho)
    
    # print(f"Fecha Pedido: {fecha_pedido.strftime('%d-%m-%Y')}")
    # print(f"Fecha Despacho: {fecha_despacho.strftime('%d-%m-%Y')}")
    # print(f"Fecha Entrega: {fecha_entrega.strftime('%d-%m-%Y') if fecha_entrega else 'N/A'}")
