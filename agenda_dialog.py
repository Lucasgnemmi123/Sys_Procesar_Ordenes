"""
Diálogo de Gestión de Agenda - UI
Creado por Lucas Gnemmi
Versión: 1.0

Interfaz gráfica para gestionar la matriz de proveedores y configuración de agenda.
"""

import tkinter as tk
from tkinter import ttk, messagebox, filedialog
from agenda_manager import AgendaManager
from datetime import datetime
from proveedor_editor import crear_editor_proveedor_mejorado


class AgendaDialog:
    """Diálogo para gestionar la configuración de agenda de proveedores"""
    
    def __init__(self, parent, theme_colors=None):
        """
        Inicializa el diálogo de agenda
        
        Args:
            parent: Ventana padre
            theme_colors: Diccionario con colores del tema
        """
        self.parent = parent
        self.manager = AgendaManager()
        
        # Colores del tema moderno (consistente con GUI principal)
        self.colors = theme_colors or {
            'bg_main': '#1a1d29',      # BG_DARK
            'bg_card': '#2d3142',      # BG_CARD  
            'bg_surface': '#242837',   # BG_SURFACE
            'fg_text': '#ffffff',      # TEXT_PRIMARY
            'fg_secondary': '#e5e7eb', # TEXT_SECONDARY
            'accent': '#00d4ff',       # PRIMARY
            'accent_hover': '#00b8d9', # PRIMARY_DARK
            'secondary': '#a78bfa',    # SECONDARY
            'button_bg': '#2d3142',    # BG_CARD
            'button_hover': '#3f4555', # BORDER
            'success': '#34d399',      # SUCCESS
            'warning': '#fbbf24',      # WARNING
            'error': '#f87171',        # ERROR
            'info': '#60a5fa'          # INFO
        }
        
        self.create_dialog()
        self.actualizar_fechas_calculadas()
        self.cargar_proveedores()
    
    def create_dialog(self):
        """Crea la ventana del diálogo"""
        # Crear un root de tkinter oculto si no existe
        # Esto evita conflictos con customtkinter
        try:
            # Intentar obtener el root existente
            root = tk._default_root
            if root is None:
                # Si no hay root, crear uno oculto
                root = tk.Tk()
                root.withdraw()
        except:
            # Si hay error, crear nuevo root oculto
            root = tk.Tk()
            root.withdraw()
        
        # Crear Toplevel independiente (compatible con customtkinter parent)
        self.dialog = tk.Toplevel()
        self.dialog.title("📅 Gestión de Agenda de Proveedores")
        
        # Adaptar altura a la pantalla del usuario
        screen_height = self.dialog.winfo_screenheight()
        window_height = int(screen_height * 0.85)  # 85% de la altura de pantalla
        window_width = 1000
        
        # Centrar ventana
        x_position = int((self.dialog.winfo_screenwidth() - window_width) / 2)
        y_position = int((screen_height - window_height) / 2)
        
        self.dialog.geometry(f"{window_width}x{window_height}+{x_position}+{y_position}")
        self.dialog.configure(bg=self.colors['bg_main'])
        self.dialog.resizable(True, True)
        
        # Header con tema moderno y botón cerrar
        header = tk.Frame(self.dialog, bg=self.colors['bg_card'], height=70)
        header.pack(fill="x")
        header.pack_propagate(False)
        
        # Frame para organizar título y botón cerrar
        header_content = tk.Frame(header, bg=self.colors['bg_card'])
        header_content.pack(fill="both", expand=True, padx=20, pady=10)
        
        tk.Label(
            header_content, 
            text="📅 GESTIÓN DE AGENDA DE PROVEEDORES", 
            font=("Segoe UI", 18, "bold"),  # Fuente más grande
            fg=self.colors['accent'],       # Color primario para el texto
            bg=self.colors['bg_card']       # Fondo del card
        ).pack(side="left")
        
        # Botón cerrar (mismo estilo que rules_dialog)
        close_btn = tk.Button(
            header_content,
            text="✕ Cerrar",
            command=self.dialog.destroy,
            bg=self.colors['error'],        # Fondo rojo
            fg='#ffffff',                   # Texto blanco
            font=("Segoe UI", 12, "bold"),
            relief='flat',
            cursor='hand2',
            padx=20,
            pady=5
        )
        close_btn.pack(side="right")
        
        # Frame principal con scroll
        main_frame = tk.Frame(self.dialog, bg=self.colors['bg_main'])
        main_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # === SECCIÓN DE CONFIGURACIÓN GENERAL ===
        config_frame = tk.LabelFrame(
            main_frame,
            text=" ⚙️ Configuración General ",
            bg=self.colors['bg_card'],
            fg=self.colors['fg_text'],
            font=('Segoe UI', 11, 'bold'),
            relief='flat',
            bd=2
        )
        config_frame.pack(fill='x', padx=5, pady=(0, 10))
        
        # Primera fila: Días de despacho
        dias_frame = tk.Frame(config_frame, bg=self.colors['bg_card'])
        dias_frame.pack(fill='x', padx=10, pady=5)
        
        tk.Label(
            dias_frame,
            text="Días para calcular Fecha de Despacho:",
            bg=self.colors['bg_card'],
            fg=self.colors['fg_text'],
            font=('Segoe UI', 10)
        ).pack(side='left', padx=(0, 10))
        
        self.dias_despacho_var = tk.StringVar(value=str(self.manager.dias_despacho))
        dias_spinbox = tk.Spinbox(
            dias_frame,
            from_=1,
            to=60,
            textvariable=self.dias_despacho_var,
            width=10,
            font=('Segoe UI', 10),
            state='readonly',
            command=self.actualizar_fechas_calculadas
        )
        dias_spinbox.pack(side='left', padx=(0, 10))
        
        tk.Button(
            dias_frame,
            text="💾 Guardar",
            command=self.guardar_dias_despacho,
            bg=self.colors['success'],
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=15,
            pady=5
        ).pack(side='left')
        
        # Botones de importación
        import_buttons_frame = tk.Frame(dias_frame, bg=self.colors['bg_card'])
        import_buttons_frame.pack(side='right')
        
        tk.Button(
            import_buttons_frame,
            text="📄 Importar Template",
            command=self.importar_desde_excel,
            bg=self.colors['secondary'],
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=12,
            pady=6
        ).pack(side='left', padx=(0, 5))
        
        tk.Button(
            import_buttons_frame,
            text="📄 Descargar Template",
            command=self.descargar_template,
            bg=self.colors['warning'],
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=12,
            pady=6
        ).pack(side='left', padx=(5, 0))
        
        # Segunda fila: Calculadora de fechas
        calc_frame = tk.Frame(config_frame, bg=self.colors['bg_card'])
        calc_frame.pack(fill='x', padx=10, pady=5)
        
        tk.Label(
            calc_frame,
            text="📅 Fecha de Pedido:",
            bg=self.colors['bg_card'],
            fg=self.colors['fg_text'],
            font=('Segoe UI', 10, 'bold')
        ).pack(side='left', padx=(0, 10))
        
        self.fecha_pedido_var = tk.StringVar(value=datetime.now().strftime("%d-%m-%Y"))
        fecha_entry = tk.Entry(
            calc_frame,
            textvariable=self.fecha_pedido_var,
            font=('Segoe UI', 10),
            width=12,
            relief='solid',
            bd=1,
            bg='#FAFAFA'
        )
        fecha_entry.pack(side='left', padx=(0, 10))
        fecha_entry.bind('<Return>', lambda e: self.actualizar_fechas_calculadas())
        
        tk.Button(
            calc_frame,
            text="🔄 Calcular Fechas",
            command=self.actualizar_fechas_calculadas,
            bg=self.colors['accent'],  # Color más visible
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=12,
            pady=6,
            activebackground='#00b8d9'  # Hover effect
        ).pack(side='left', padx=(0, 20))
        
        tk.Label(
            calc_frame,
            text="📦 Fecha Despacho:",
            bg=self.colors['bg_card'],
            fg=self.colors['fg_text'],
            font=('Segoe UI', 10, 'bold')
        ).pack(side='left', padx=(0, 10))
        
        self.fecha_despacho_label = tk.Label(
            calc_frame,
            text="--/--/----",
            bg=self.colors['bg_card'],
            fg=self.colors['accent'],
            font=('Segoe UI', 10, 'bold')
        )
        self.fecha_despacho_label.pack(side='left')
        
        # === SECCIÓN DE LISTA DE PROVEEDORES ===
        lista_frame = tk.LabelFrame(
            main_frame,
            text=" 📋 Proveedores Configurados ",
            bg=self.colors['bg_card'],
            fg=self.colors['fg_text'],
            font=('Segoe UI', 11, 'bold'),
            relief='flat',
            bd=2
        )
        lista_frame.pack(fill='both', expand=True, padx=5, pady=(0, 10))
        
        # Tabla de proveedores
        tabla_frame = tk.Frame(lista_frame, bg=self.colors['bg_card'])
        tabla_frame.pack(fill='both', expand=True, padx=10, pady=10)
        
        # Scrollbars
        scrollbar_y = tk.Scrollbar(tabla_frame, orient='vertical')
        scrollbar_y.pack(side='right', fill='y')
        
        scrollbar_x = tk.Scrollbar(tabla_frame, orient='horizontal')
        scrollbar_x.pack(side='bottom', fill='x')
        
        # Configurar estilo del Treeview con bordes negros
        style = ttk.Style()
        style.theme_use('default')
        style.configure('Bordered.Treeview',
            background='white',
            foreground='black',
            fieldbackground='white',
            bordercolor='black',
            borderwidth=1,
            rowheight=28,
            font=('Segoe UI', 9)
        )
        style.configure('Bordered.Treeview.Heading',
            background='#E3F2FD',
            foreground='#1976D2',
            borderwidth=1,
            font=('Segoe UI', 9, 'bold'),
            relief='solid'
        )
        style.map('Bordered.Treeview',
            background=[('selected', '#BBDEFB')],
            foreground=[('selected', 'black')]
        )
        
        # Treeview
        columnas = ('Código', 'Nombre', 'LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'D-2', 'Fecha Entrega')
        self.tree = ttk.Treeview(
            tabla_frame,
            columns=columnas,
            show='headings',
            yscrollcommand=scrollbar_y.set,
            xscrollcommand=scrollbar_x.set,
            height=15,
            style='Bordered.Treeview'
        )
        
        scrollbar_y.config(command=self.tree.yview)
        scrollbar_x.config(command=self.tree.xview)
        
        # Configurar columnas con bordes visibles
        anchos = {'Código': 80, 'Nombre': 250, 'LUN': 45, 'MAR': 45, 'MIE': 45, 'JUE': 45, 'VIE': 45, 'SAB': 45, 'D-2': 45, 'Fecha Entrega': 120}
        for col in columnas:
            self.tree.heading(col, text=col)
            self.tree.column(col, width=anchos.get(col, 100), anchor='center' if col not in ['Nombre'] else 'w', minwidth=anchos.get(col, 100))
        
        # Configurar tags de filas alternadas para mejorar la separación visual
        self.tree.tag_configure('oddrow', background='#F5F5F5')
        self.tree.tag_configure('evenrow', background='white')
        # Tag especial para proveedores con fecha manual (amarillo claro)
        self.tree.tag_configure('fecha_manual', background='#FFEB3B')
        
        self.tree.pack(fill='both', expand=True)
        
        # Agregar indicador visual de que las columnas de días son clickeables
        tk.Label(
            tabla_frame,
            text="💡 Tip: Haz clic en las columnas de días (LUN-SAB, D-2) para cambiar sus valores",
            bg=self.colors['bg_main'],
            fg='#666666',
            font=('Segoe UI', 9, 'italic'),
            anchor='w'
        ).pack(fill='x', pady=(5, 0))
        
        # Agregar evento de clic para editar días
        self.tree.bind('<Button-1>', self.on_tree_click)
        self.tree.bind('<Motion>', self.on_tree_motion)
        
        # === SECCIÓN DE BOTONES ===
        botones_frame = tk.Frame(main_frame, bg=self.colors['bg_main'])
        botones_frame.pack(fill='x', padx=5, pady=(0, 5))
        
        tk.Button(
            botones_frame,
            text="➕ Agregar Proveedor",
            command=self.agregar_proveedor,
            bg=self.colors['success'],
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=15,
            pady=6
        ).pack(side='left', padx=5)
        
        tk.Button(
            botones_frame,
            text="✏️ Editar Seleccionado",
            command=self.editar_proveedor,
            bg=self.colors['info'],
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=15,
            pady=6
        ).pack(side='left', padx=5)
        
        tk.Button(
            botones_frame,
            text="🗑️ Eliminar Seleccionado",
            command=self.eliminar_proveedor,
            bg=self.colors['accent'],
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=15,
            pady=6
        ).pack(side='left', padx=5)
        
        # Botón para eliminar toda la agenda
        tk.Button(
            botones_frame,
            text="🗑️ Eliminar Toda la Agenda",
            command=self.eliminar_toda_agenda,
            bg=self.colors['error'],
            fg='white',
            font=('Segoe UI', 9, 'bold'),
            relief='flat',
            cursor='hand2',
            padx=15,
            pady=6
        ).pack(side='left', padx=5)
    
    def cargar_proveedores(self):
        """Carga los proveedores en la tabla"""
        # Limpiar tabla
        for item in self.tree.get_children():
            self.tree.delete(item)
        
        # Obtener fecha de despacho si está calculada
        fecha_despacho = None
        try:
            fecha_pedido_str = self.fecha_pedido_var.get()
            fecha_pedido = datetime.strptime(fecha_pedido_str, "%d-%m-%Y")
            fecha_despacho = self.manager.calcular_fecha_despacho(fecha_pedido)
        except:
            pass
        
        # Cargar proveedores
        proveedores = self.manager.obtener_todos_proveedores()
        for codigo, datos in proveedores.items():
            def formato_dia(valor):
                """Convierte valor a símbolo: 1='✓', 0='○', None=''"""
                if valor == 1 or valor is True:
                    return '✓'
                elif valor == 0:
                    return '○'
                else:
                    return ''
            
            # Calcular fecha de entrega para este proveedor
            fecha_entrega_str = ''
            if fecha_despacho:
                fecha_entrega = self.manager.calcular_fecha_entrega(codigo, fecha_despacho)
                if fecha_entrega:
                    fecha_entrega_str = fecha_entrega.strftime("%d-%m-%Y")
                elif datos.get('fecha_manual'):
                    fecha_entrega_str = datos.get('fecha_manual') + ' (Manual)'
            elif datos.get('fecha_manual'):
                fecha_entrega_str = datos.get('fecha_manual') + ' (Manual)'
            
            valores = (
                codigo,
                datos.get('nombre', ''),
                formato_dia(datos.get('LUN')),
                formato_dia(datos.get('MAR')),
                formato_dia(datos.get('MIE')),
                formato_dia(datos.get('JUE')),
                formato_dia(datos.get('VIE')),
                formato_dia(datos.get('SAB')),
                formato_dia(datos.get('D-2')),
                fecha_entrega_str
            )
            
            # Determinar tag según los valores de días
            dias_valores = [
                datos.get('LUN'),
                datos.get('MAR'),
                datos.get('MIE'),
                datos.get('JUE'),
                datos.get('VIE'),
                datos.get('SAB'),
                datos.get('D-2')
            ]
            
            # Contar tipos de valores para determinar el color predominante
            tiene_si = any(v == 1 or v is True for v in dias_valores)
            tiene_no = any(v == 0 for v in dias_valores)
            todos_none = all(v is None for v in dias_valores)
            
            # Determinar tag de fila: amarillo si tiene fecha manual, sino alternado
            if datos.get('fecha_manual'):
                tag_fila = 'fecha_manual'
            else:
                num_items = len(self.tree.get_children())
                tag_fila = 'evenrow' if num_items % 2 == 0 else 'oddrow'
            
            self.tree.insert('', 'end', values=valores, tags=(tag_fila,))
    
    def guardar_dias_despacho(self):
        """Guarda los días de despacho configurados"""
        try:
            dias = int(self.dias_despacho_var.get())
            if dias < 1:
                messagebox.showerror("Error", "Los días deben ser mayor a 0", parent=self.dialog)
                return
            
            self.manager.dias_despacho = dias
            self.manager.guardar_configuracion()
            self.actualizar_fechas_calculadas()
            messagebox.showinfo("✅ Guardado", f"Días de despacho configurados: {dias}", parent=self.dialog)
        except ValueError:
            messagebox.showerror("Error", "Ingrese un número válido", parent=self.dialog)
    
    def actualizar_fechas_calculadas(self):
        """Actualiza las fechas de despacho y entrega calculadas"""
        try:
            # Validar y parsear fecha de pedido
            fecha_pedido_str = self.fecha_pedido_var.get()
            fecha_pedido = datetime.strptime(fecha_pedido_str, "%d-%m-%Y")
            
            # Actualizar días de despacho si cambió
            try:
                dias = int(self.dias_despacho_var.get())
                if dias > 0:
                    self.manager.dias_despacho = dias
            except:
                pass
            
            # Calcular fecha de despacho
            fecha_despacho = self.manager.calcular_fecha_despacho(fecha_pedido)
            self.fecha_despacho_label.config(text=fecha_despacho.strftime("%d-%m-%Y"))
            
            # Recargar proveedores para mostrar fechas de entrega calculadas
            self.cargar_proveedores()
            
        except ValueError:
            self.fecha_despacho_label.config(text="Formato inválido")
            messagebox.showerror("Error", "Formato de fecha inválido. Use dd-mm-yyyy", parent=self.dialog)
        except Exception as e:
            self.fecha_despacho_label.config(text="Error")
            messagebox.showerror("Error", f"Error calculando fechas: {e}", parent=self.dialog)
    
    def on_tree_click(self, event):
        """Maneja clic en la tabla para editar días directamente"""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            return
        
        # Identificar columna y fila
        column_id = self.tree.identify_column(event.x)
        item_id = self.tree.identify_row(event.y)
        
        if not item_id:
            return
        
        # Obtener índice de columna (empiezan en #1)
        col_index = int(column_id.replace('#', '')) - 1
        columnas = ('Código', 'Nombre', 'LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'D-2', 'Fecha Entrega')
        
        if col_index < 0 or col_index >= len(columnas):
            return
        
        col_name = columnas[col_index]
        
        # Solo permitir edición en columnas de días
        dias_editables = ['LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'D-2']
        if col_name not in dias_editables:
            return
        
        # Obtener datos del proveedor
        valores = self.tree.item(item_id)['values']
        codigo_prov = valores[0]
        
        # Obtener valor actual
        valor_actual_str = valores[col_index]
        
        # Convertir símbolo a valor: ✓=1, ○=0, vacío=None
        if valor_actual_str == '✓':
            valor_actual = 1
        elif valor_actual_str == '○':
            valor_actual = 0
        else:
            valor_actual = None
        
        # Ciclar al siguiente estado: None → 1 → 0 → None
        if valor_actual is None:
            nuevo_valor = 1
        elif valor_actual == 1:
            nuevo_valor = 0
        else:
            nuevo_valor = None
        
        # Actualizar en el manager
        proveedor_data = self.manager.obtener_proveedor(codigo_prov)
        if proveedor_data:
            proveedor_data[col_name] = nuevo_valor
            self.manager.agregar_proveedor(
                codigo_prov,
                proveedor_data['nombre'],
                {dia: proveedor_data.get(dia) for dia in ['LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB']},
                proveedor_data.get('D-2'),
                proveedor_data.get('fecha_manual')
            )
            
            # Recargar la tabla para reflejar el cambio
            self.cargar_proveedores()
    
    def on_tree_motion(self, event):
        """Cambia el cursor cuando está sobre columnas editables"""
        region = self.tree.identify_region(event.x, event.y)
        if region != "cell":
            self.tree.config(cursor="")
            return
        
        column_id = self.tree.identify_column(event.x)
        col_index = int(column_id.replace('#', '')) - 1
        columnas = ('Código', 'Nombre', 'LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'D-2', 'Fecha Entrega')
        
        if col_index >= 0 and col_index < len(columnas):
            col_name = columnas[col_index]
            dias_editables = ['LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'D-2']
            if col_name in dias_editables:
                self.tree.config(cursor="hand2")
            else:
                self.tree.config(cursor="")
    
    def agregar_proveedor(self):
        """Abre diálogo para agregar un nuevo proveedor"""
        self.abrir_editor_proveedor()
    
    def editar_proveedor(self):
        """Edita el proveedor seleccionado"""
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning("Selección", "Seleccione un proveedor para editar", parent=self.dialog)
            return
        
        item = self.tree.item(seleccion[0])
        codigo = item['values'][0]
        self.abrir_editor_proveedor(codigo)
    
    def eliminar_proveedor(self):
        """Elimina el proveedor seleccionado"""
        seleccion = self.tree.selection()
        if not seleccion:
            messagebox.showwarning("Selección", "Seleccione un proveedor para eliminar", parent=self.dialog)
            return
        
        item = self.tree.item(seleccion[0])
        codigo = item['values'][0]
        nombre = item['values'][1]
        
        if messagebox.askyesno("Confirmar", f"¿Eliminar proveedor {codigo} - {nombre}?", parent=self.dialog):
            self.manager.eliminar_proveedor(codigo)
            self.cargar_proveedores()
            messagebox.showinfo("✅ Eliminado", "Proveedor eliminado correctamente", parent=self.dialog)
    
    def abrir_editor_proveedor(self, codigo_editar=None):
        """Abre diálogo para agregar/editar proveedor usando el editor mejorado"""
        crear_editor_proveedor_mejorado(
            dialog=self.dialog,
            manager=self.manager,
            colors=self.colors,
            codigo_editar=codigo_editar,
            callback_actualizar=self.cargar_proveedores
        )
    
    def importar_desde_excel(self):
        """Importa proveedores desde un archivo Agenda.xlsm"""
        archivo = filedialog.askopenfilename(
            title="Seleccionar Agenda.xlsm",
            filetypes=[("Excel Macro", "*.xlsm"), ("Todos los archivos", "*.*")]
        )
        
        if archivo:
            if self.manager.importar_desde_excel(archivo):
                self.cargar_proveedores()
                messagebox.showinfo("✅ Importado", "Proveedores importados correctamente desde Excel", parent=self.dialog)
            else:
                messagebox.showerror("Error", "No se pudo importar la configuración desde Excel", parent=self.dialog)
    
    def descargar_template(self):
        """Descarga un template de Excel con la estructura esperada"""
        import pandas as pd
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment
        
        # Pedir ubicación de guardado
        archivo = filedialog.asksaveasfilename(
            title="Guardar Template",
            defaultextension=".xlsx",
            filetypes=[("Excel files", "*.xlsx"), ("All files", "*.*")],
            initialfile="Agenda_Template.xlsx"
        )
        
        if not archivo:
            return
        
        try:
            # Crear workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Matriz"
            
            # Encabezado principal (fila 1)
            ws['A1'] = "AGENDA DE PROVEEDORES"
            ws.merge_cells('A1:K1')
            ws['A1'].font = Font(size=14, bold=True)
            ws['A1'].fill = PatternFill(start_color="1F4E78", end_color="1F4E78", fill_type="solid")
            ws['A1'].font = Font(size=14, bold=True, color="FFFFFF")
            ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
            
            # Encabezados de columnas (fila 2)
            headers = ['CODIGO', 'PROVEEDOR', '', 'LUN', 'MAR', 'MIE', 'JUE', 'VIE', 'SAB', 'D+2']
            for col, header in enumerate(headers, start=1):
                cell = ws.cell(row=2, column=col)
                cell.value = header
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
                cell.alignment = Alignment(horizontal='center', vertical='center')
            
            # Datos de ejemplo (fila 3 en adelante)
            ejemplos = [
                ['PROV001', 'PROVEEDOR EJEMPLO 1', '', 1, 0, 1, 0, 1, 0, 0],
                ['PROV002', 'PROVEEDOR EJEMPLO 2', '', 0, 1, 0, 1, 0, 0, 1],
                ['PROV003', 'PROVEEDOR EJEMPLO 3', '', 1, 1, 1, 1, 1, 0, 0]
            ]
            
            for row_idx, ejemplo in enumerate(ejemplos, start=3):
                for col_idx, valor in enumerate(ejemplo, start=1):
                    ws.cell(row=row_idx, column=col_idx, value=valor)
            
            # Ajustar anchos de columnas
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 30
            ws.column_dimensions['C'].width = 3
            for col in ['D', 'E', 'F', 'G', 'H', 'I', 'J']:
                ws.column_dimensions[col].width = 6
            
            # Guardar
            wb.save(archivo)
            
            messagebox.showinfo(
                "✅ Template Descargado",
                f"Template guardado en:\n{archivo}\n\n"
                "Estructura:\n"
                "- CODIGO: Código del proveedor\n"
                "- PROVEEDOR: Nombre del proveedor\n"
                "- Columna vacía (separador)\n"
                "- LUN-SAB: Días de entrega (1=entrega, 0=no entrega)\n"
                "- D+2: Entrega D+2 (1=sí, 0=no)",
                parent=self.dialog
            )
        except Exception as e:
            messagebox.showerror(
                "❌ Error",
                f"No se pudo crear el template:\n{str(e)}",
                parent=self.dialog
            )

    def eliminar_toda_agenda(self):
        """Elimina toda la agenda de proveedores"""
        if messagebox.askyesno(
            "🗑️ Confirmar Eliminación", 
            "\u00bfEstá seguro de que desea eliminar TODA la agenda de proveedores?\n\nEsta acción NO se puede deshacer.",
            parent=self.dialog
        ):
            try:
                # Limpiar los proveedores (el atributo correcto es 'proveedores', no 'agenda')
                self.manager.proveedores.clear()  # Método más seguro
                
                # Guardar cambios
                self.manager.guardar_configuracion()
                
                # Actualizar la tabla
                self.cargar_proveedores()
                messagebox.showinfo(
                    "✅ Agenda Eliminada", 
                    "Toda la agenda de proveedores ha sido eliminada correctamente.",
                    parent=self.dialog
                )
            except Exception as e:
                messagebox.showerror(
                    "Error", 
                    f"No se pudo eliminar la agenda: {str(e)}", 
                    parent=self.dialog
                )


def abrir_dialogo_agenda(parent, theme_colors=None):
    """Función de utilidad para abrir el diálogo"""
    AgendaDialog(parent, theme_colors)


if __name__ == "__main__":
    # Prueba del diálogo
    root = tk.Tk()
    root.withdraw()
    abrir_dialogo_agenda(root)
    root.mainloop()
