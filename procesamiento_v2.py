"""
DHL Order Processing System - Processing Module
Created by Lucas Gnemmi
Version: 2.0

Módulo optimizado para el procesamiento de órdenes de DHL.
Incluye extracción de PDFs, validación de SKUs, mapeo de proveedores y formateo de Excel.
"""

import os
import sys

# Agregar directorio del script al path
script_dir = os.path.dirname(os.path.abspath(__file__))
if script_dir not in sys.path:
    sys.path.insert(0, script_dir)

# Agregar carpeta libs al path
libs_path = os.path.join(script_dir, 'libs')
if os.path.exists(libs_path) and libs_path not in sys.path:
    sys.path.insert(0, libs_path)

import re
# import fitz  # PyMuPDF - NO NECESARIO, ya no procesamos PDFs
import pandas as pd
import xlwings as xw
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

# --- Utilidades y funciones auxiliares optimizadas ---

def clean_qty(qtext):
    """
    Limpia y convierte texto de cantidad a número (float)
    Maneja formatos europeos con puntos y comas
    Mantiene las cantidades decimales exactamente como vienen
    """
    q = str(qtext).strip()
    
    # Manejar formato europeo (1.234,56)
    if '.' in q and ',' in q:
        q = q.replace('.', '').replace(',', '.')
    else:
        q = q.replace(',', '.')
    
    try:
        return float(q)  # Mantener valor original sin redondear
    except (ValueError, TypeError):
        print(f"⚠️ Warning: Cannot convert quantity '{qtext}' to number")
        return q

def limpiar_nombre_lugar(nombre):
    """
    Limpia NOMBRE_LUGAR eliminando todas las variaciones de BOD. y ENAP
    para hacer la observación más corta y clara
    """
    if pd.isna(nombre):
        return ""
    nombre = str(nombre)
    
    # Eliminar todas las variaciones de BOD. (con y sin punto)
    nombre = nombre.replace("BOD.", "")
    nombre = nombre.replace("BOD ", "")
    nombre = nombre.replace("BOD", "")
    
    # Eliminar todas las variaciones de ENAP MAGALLANES
    nombre = nombre.replace("ENAP MAGALLANES", "")
    nombre = nombre.replace("ENAP MAG.", "")
    nombre = nombre.replace("ENAP MAG ", "")  
    nombre = nombre.replace("ENAP MAG", "")
    nombre = nombre.replace("ENAP MAGA", "")
    nombre = nombre.replace("ENAP MA", "")
    nombre = nombre.replace("ENAP", "")
    
    # Eliminar MAGALLANES suelto
    nombre = nombre.replace("MAGALLANES", "")
    
    # Limpiar espacios múltiples y al inicio/final
    nombre = ' '.join(nombre.split())  # Elimina espacios múltiples
    
    return nombre.strip()

def extract_centro_costo_y_nombre(text):
    """
    Extrae centro de costo y nombre del lugar desde el texto del PDF
    Mejorado con múltiples patrones de búsqueda
    """
    # Patrón principal
    patterns = [
        r'Nombre\s*:\s*(\d{3,6})\s+(.*)',
        r'Centro\s*:\s*(\d{3,6})\s+(.*)',
        r'Código\s*:\s*(\d{3,6})\s+(.*)'
    ]
    
    for pattern in patterns:
        match = re.search(pattern, text, flags=re.IGNORECASE)
        if match:
            centro_costo = match.group(1).strip()
            nombre = match.group(2).strip()
            # Limpiar el nombre de caracteres extra
            nombre = re.sub(r'[^\w\s\-\.]', '', nombre).strip()
            return centro_costo, nombre
    
    print("⚠️ Warning: Could not extract center cost and name from PDF")
    return "", ""

def extract_items_from_text(text):
    """
    Extrae items (SKU y cantidad) del texto del PDF
    Mejorado con múltiples patrones y validación
    """
    # Patrón principal mejorado
    patterns = [
        r'^\s*\d+\s+\d+\s+\d{2}/\d{2}/\d{2}\s+(A\d{4,6})\s+(?:A\d{4,6}\s+)?([\d\.\,]+)',
        r'^\s*\d+\s+\d+\s+\d{4}-\d{2}-\d{2}\s+(A\d{4,6})\s+(?:A\d{4,6}\s+)?([\d\.\,]+)',
        r'(A\d{4,6})\s+([\d\.\,]+)\s*$'
    ]
    
    items = []
    
    for pattern in patterns:
        matches = re.findall(pattern, text, re.MULTILINE)
        for sku, qty in matches:
            # Validar SKU
            if sku.startswith('A') and len(sku) >= 5:
                items.append({"SKU": sku.upper(), "CANTIDAD_RAW": qty})
    
    # Eliminar duplicados manteniendo el orden
    seen = set()
    unique_items = []
    for item in items:
        identifier = (item["SKU"], item["CANTIDAD_RAW"])
        if identifier not in seen:
            seen.add(identifier)
            unique_items.append(item)
    
    return unique_items

# --- Procesamiento principal de PDFs optimizado ---

def procesar_pdfs(ordenes_dir):
    """
    Procesa archivo Excel en la carpeta de órdenes (anteriormente procesaba PDFs)
    Lee datos desde Excel con columnas: LOCAL_ENTREGA_CTRPED, DESCR_CEN_CADCEN, COD_MAT_PEDCOM, QTDE_PEDIDA_PEDCOM
    """
    if not os.path.exists(ordenes_dir):
        raise FileNotFoundError(f"❌ Orders folder not found: {ordenes_dir}")

    print(f"📂 Processing Excel files from: {ordenes_dir}")
    
    rows = []
    archivos_procesados = 0
    archivos_con_errores = 0
    
    # Buscar archivos Excel en lugar de PDFs
    excel_files = [f for f in sorted(os.listdir(ordenes_dir)) if f.lower().endswith(('.xlsx', '.xls'))]
    
    if not excel_files:
        print("⚠️ No Excel files found in orders folder")
        return pd.DataFrame(columns=["LOCAL", "SKU", "CANTIDAD", "CENTRO_COSTO", "NOMBRE_LUGAR", "_SRC_FILE"])
    
    print(f"📄 Found {len(excel_files)} Excel files to process")
    
    for fname in excel_files:
        try:
            path = os.path.join(ordenes_dir, fname)
            print(f"📖 Processing: {fname}")
            
            # Leer archivo Excel
            try:
                df_excel = pd.read_excel(path, dtype=str)
            except Exception as e:
                try:
                    df_excel = pd.read_excel(path, engine='openpyxl', dtype=str)
                except Exception as e2:
                    print(f"❌ Error reading Excel file {fname}: {e2}")
                    archivos_con_errores += 1
                    continue
            
            # Limpiar nombres de columnas
            df_excel.columns = df_excel.columns.str.strip()
            
            # Verificar que existan las columnas requeridas
            required_columns = ['LOCAL_ENTREGA_CTRPED', 'DESCR_CEN_CADCEN', 'COD_MAT_PEDCOM', 'QTDE_PEDIDA_PEDCOM']
            missing_columns = [col for col in required_columns if col not in df_excel.columns]
            
            if missing_columns:
                print(f"❌ Missing columns in {fname}: {missing_columns}")
                print(f"📋 Available columns: {list(df_excel.columns)}")
                archivos_con_errores += 1
                continue
            
            # Procesar cada fila del Excel
            items_procesados = 0
            items_rechazados = 0
            print(f"📋 Total rows in Excel: {len(df_excel)}")
            
            for idx, row in df_excel.iterrows():
                try:
                    # Extraer datos de las columnas
                    centro_costo = str(row['LOCAL_ENTREGA_CTRPED']).strip()
                    nombre_lugar = str(row['DESCR_CEN_CADCEN']).strip()
                    sku = str(row['COD_MAT_PEDCOM']).strip().upper()
                    cantidad_raw = str(row['QTDE_PEDIDA_PEDCOM']).strip()
                    
                    # Limpiar y validar cantidad
                    qty = clean_qty(cantidad_raw)
                    
                    # Validaciones más flexibles
                    sku_valido = sku and sku != 'NAN' and sku.lower() != 'nan' and len(sku) > 0
                    centro_valido = centro_costo and centro_costo != 'nan' and centro_costo.lower() != 'nan'
                    cantidad_valida = isinstance(qty, (int, float)) and qty > 0
                    
                    if sku_valido and cantidad_valida:
                        # Usar valores por defecto si faltan datos opcionales
                        if not centro_valido:
                            centro_costo = "UNKNOWN"
                        if not nombre_lugar or nombre_lugar == 'nan' or nombre_lugar.lower() == 'nan':
                            nombre_lugar = "UNKNOWN"
                        
                        # Convertir cantidad a float para mantener decimales
                        cantidad_final = float(qty)
                        
                        rows.append({
                            "LOCAL": "30797",
                            "SKU": sku,
                            "CANTIDAD": cantidad_final,  # Mantener como float
                            "CENTRO_COSTO": centro_costo,
                            "NOMBRE_LUGAR": nombre_lugar,
                            "_SRC_FILE": fname
                        })
                        items_procesados += 1
                    else:
                        items_rechazados += 1
                        razon = []
                        if not sku_valido:
                            razon.append(f"SKU invalid: '{sku}'")
                        if not cantidad_valida:
                            razon.append(f"QTY invalid: '{cantidad_raw}' -> {qty}")
                        
                        print(f"⚠️ Row {idx+1} rejected: {'; '.join(razon)}")
                        
                except Exception as e:
                    items_rechazados += 1
                    print(f"⚠️ Error processing row {idx+1}: {e}")
            
            print(f"📊 Processing results for {fname}:")
            print(f"   • Total rows: {len(df_excel)}")
            print(f"   • Valid items: {items_procesados}")
            print(f"   • Rejected items: {items_rechazados}")
            
            if items_procesados > 0:
                print(f"✅ {fname}: {items_procesados} items extracted")
                archivos_procesados += 1
            else:
                print(f"❌ {fname}: No valid items found")
                archivos_con_errores += 1
                
        except Exception as e:
            print(f"❌ Error processing {fname}: {e}")
            archivos_con_errores += 1

    print(f"📊 Processing summary:")
    print(f"   • Files processed successfully: {archivos_procesados}")
    print(f"   • Files with errors: {archivos_con_errores}")
    print(f"   • Total records extracted: {len(rows)}")

    return pd.DataFrame(rows).reset_index(drop=True)

# --- Validación de SKUs optimizada ---

def validar_skus_items(df, products_manager=None):
    """
    Valida que los SKUs estén en la lista maestra de productos (compra calzada)
    Ahora usa ProductsManager en lugar de Items.xlsx
    """
    from products_manager import ProductsManager
    
    df = df.copy()
    df_err = pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
    warnings = []

    try:
        # Inicializar ProductsManager si no se proporciona
        if products_manager is None:
            products_manager = ProductsManager()
        
        print(f"🔍 Validating SKUs against Products Master List...")
        
        # Obtener todos los SKUs válidos
        skus_validos = products_manager.get_all_skus()
        
        if not skus_validos:
            warnings.append(f"⚠️ No products found in master list - validation skipped")
            warnings.append(f"💡 Use Products Manager to add products")
            return df, df_err, warnings
        
        warnings.append(f"✅ Products loaded: {len(skus_validos)} valid SKUs found")
        
        # Validar cada SKU
        df_skus = df["SKU"].str.strip().str.upper()
        mask_no_en_items = ~df_skus.isin(skus_validos)
        
        if mask_no_en_items.any():
            df_err_items = df.loc[mask_no_en_items].copy()
            df_err_items["OBSERVACION"] = (
                df_err_items["CENTRO_COSTO"].fillna("") + 
                "//Falta Producto en Maestra C.Calzada//" + 
                df_err_items["NOMBRE_LUGAR"].fillna("")
            )
            df_err = pd.concat([df_err, df_err_items], ignore_index=True)
            
            # Mostrar algunos ejemplos de SKUs no encontrados
            skus_no_encontrados = df_skus[mask_no_en_items].unique()[:5]
            warnings.append(f"⚠️ {len(df_err_items)} SKUs faltan en Maestra C.Calzada")
            warnings.append(f"📋 Examples: {list(skus_no_encontrados)}")
            warnings.append(f"💡 Add missing SKUs via Products Manager")
        
        # Retornar solo los válidos
        df_valid = df.loc[~mask_no_en_items].copy()
        warnings.append(f"✅ Valid SKUs: {len(df_valid)}")
        
    except Exception as e:
        warnings.append(f"❌ Error validating products: {e}")
        df_valid = df.copy()  # Si hay error, continuar sin validación
    
    return df_valid, df_err, warnings

# --- Mapeo de proveedores optimizado ---

def mapear_proveedor_por_sku(df, full_xlsx, region="099", apply_rules=True):
    """
    Mapea proveedores por SKU desde Full.xlsx
    Versión optimizada con mejor manejo de datos y logging
    
    Args:
        df: DataFrame con datos a procesar
        full_xlsx: Ruta al archivo Full.xlsx
        region: Región a filtrar (default "099")
        apply_rules: Si True, aplica reglas especiales (default True)
    """
    print(f"🔍 Mapping suppliers from: {full_xlsx}")
    print(f"📍 Using region: {region}")
    
    df = df.copy()
    df_err = pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
    warnings = []
    
    # Cargar reglas especiales si están habilitadas
    rules_manager = None
    if apply_rules:
        try:
            from rules_manager import RulesManager
            rules_manager = RulesManager()
            stats = rules_manager.get_stats()
            if stats['active_local_rules'] > 0 or stats['active_stock_blocks'] > 0:
                warnings.append(f"⚙️ Special rules loaded: {stats['active_local_rules']} LOCAL rules, {stats['active_stock_blocks']} stock blocks")
                print(f"⚙️ Applying special rules: {stats['active_local_rules']} LOCAL rules, {stats['active_stock_blocks']} stock blocks")
            else:
                rules_manager = None  # No hay reglas activas
        except Exception as e:
            warnings.append(f"⚠️ Could not load special rules: {e}")
            rules_manager = None

    try:
        # Verificar que el archivo exista
        if not os.path.exists(full_xlsx):
            warnings.append(f"❌ Full.xlsx not found: {full_xlsx}")
            return df, df_err, warnings

        # Leer Full.xlsx
        try:
            df_full = pd.read_excel(full_xlsx, dtype=str)
        except Exception as e:
            try:
                df_full = pd.read_excel(full_xlsx, engine='openpyxl', dtype=str)
            except Exception as e2:
                warnings.append(f"❌ Error reading Full.xlsx: {e2}")
                return df, df_err, warnings
        
        # Limpiar nombres de columnas
        df_full.columns = df_full.columns.str.strip()
        warnings.append(f"📊 Full.xlsx loaded: {len(df_full)} records")
        warnings.append(f"📋 Columns: {list(df_full.columns)}")
        
        # Buscar columnas necesarias
        col_sku_full = None
        col_proveedor = None
        
        # Buscar columna SKU
        for col in df_full.columns:
            if any(keyword in col.lower() for keyword in ['sku', 'codigo', 'code', 'artículo', 'articulo']):
                col_sku_full = col
                break
        
        # Buscar columna proveedor
        for col in df_full.columns:
            if any(keyword in col.lower() for keyword in ['proveedor', 'supplier', 'vendor']):
                col_proveedor = col
                break
        
        if not col_sku_full:
            warnings.append(f"❌ SKU column not found in Full.xlsx")
            return df, df_err, warnings
            
        if not col_proveedor:
            warnings.append(f"❌ Supplier column not found in Full.xlsx")
            return df, df_err, warnings
        
        warnings.append(f"✅ Using SKU column: {col_sku_full}")
        warnings.append(f"✅ Using supplier column: {col_proveedor}")
        
        # Filtrar por región - buscar columna específicamente
        col_region = None
        possible_region_columns = ['Región', 'Region', 'region', 'REGION', 'zona', 'Zona', 'ZONA']
        
        for col_name in possible_region_columns:
            if col_name in df_full.columns:
                col_region = col_name
                break
        
        if not col_region:
            # Buscar por coincidencia parcial
            for col in df_full.columns:
                if any(keyword in col.lower() for keyword in ['region', 'zona', 'area']):
                    col_region = col
                    break
        
        if col_region:
            # Filtrar por región específica
            mask_region = df_full[col_region].astype(str).str.strip() == str(region)
            df_region = df_full[mask_region].copy()
            
            # Mostrar información de filtrado
            total_regions = df_full[col_region].astype(str).str.strip().unique()
            warnings.append(f"🌍 Found region column: {col_region}")
            warnings.append(f"📊 Available regions: {sorted(total_regions)}")
            warnings.append(f"� Filtered by region '{region}': {len(df_region)} records (from {len(df_full)} total)")
            
            if len(df_region) == 0:
                warnings.append(f"⚠️ No records found for region '{region}', using all records")
                df_region = df_full.copy()
        else:
            df_region = df_full.copy()
            warnings.append(f"⚠️ No region column found, using all records")
        
        # Crear diccionario de mapeo SKU -> Lista de Proveedores (puede haber múltiples)
        df_region_clean = df_region.dropna(subset=[col_sku_full, col_proveedor])
        sku_to_proveedores = {}  # SKU -> lista de proveedores disponibles
        
        for _, row in df_region_clean.iterrows():
            sku = str(row[col_sku_full]).strip().upper()
            proveedor = str(row[col_proveedor]).strip()
            # Normalizar código de proveedor eliminando .0 si existe
            proveedor = proveedor.replace('.0', '') if proveedor.endswith('.0') else proveedor
            
            if sku and proveedor and sku != 'NAN' and proveedor != 'NAN':
                if sku not in sku_to_proveedores:
                    sku_to_proveedores[sku] = []
                if proveedor not in sku_to_proveedores[sku]:
                    sku_to_proveedores[sku].append(proveedor)
        
        warnings.append(f"📋 Created mapping for {len(sku_to_proveedores)} SKUs")
        
        # Mostrar SKUs con múltiples proveedores
        multi_prov = {sku: provs for sku, provs in sku_to_proveedores.items() if len(provs) > 1}
        if multi_prov:
            warnings.append(f"🔀 {len(multi_prov)} SKUs have multiple suppliers:")
            for sku, provs in list(multi_prov.items())[:3]:  # Mostrar solo 3 ejemplos
                warnings.append(f"   • {sku}: {len(provs)} suppliers → {provs}")
        
        # Aplicar mapeo con reglas especiales
        df_mapped = []
        df_errors = []
        reglas_aplicadas_local = 0
        reglas_aplicadas_bloqueo = 0
        
        for _, row in df.iterrows():
            sku = str(row["SKU"]).strip().upper()
            local = str(row.get("CENTRO_COSTO", "")).strip()
            
            # Verificar si el SKU existe en el mapeo
            if sku not in sku_to_proveedores:
                # SKU no encontrado en Full.xlsx
                row_error = row.copy()
                row_error["OBSERVACION"] = (
                    str(row.get("CENTRO_COSTO", "")) + 
                    "//No tiene Precio//" + 
                    str(row.get("NOMBRE_LUGAR", ""))
                )
                df_errors.append(row_error)
                continue
            
            # Obtener proveedores disponibles para este SKU
            proveedores_disponibles = sku_to_proveedores[sku].copy()
            
            # REGLA 1: Verificar si hay regla de LOCAL + SKU → Proveedor forzado
            # MÁXIMA PRIORIDAD - Esta regla sobrescribe todo lo demás
            proveedor_forzado = None
            tiene_regla_especial = False
            if rules_manager:
                proveedor_forzado = rules_manager.get_proveedor_for_local_sku(local, sku)
                if proveedor_forzado:
                    tiene_regla_especial = True
                    # Normalizar código de proveedor forzado (eliminar .0 si existe)
                    proveedor_forzado_norm = proveedor_forzado.replace('.0', '') if proveedor_forzado.endswith('.0') else proveedor_forzado
                    
                    # Verificar si el proveedor forzado existe en los disponibles
                    if proveedor_forzado_norm in proveedores_disponibles:
                        # FORZAR este proveedor ignorando todo lo demás
                        proveedores_disponibles = [proveedor_forzado_norm]
                        reglas_aplicadas_local += 1
                        print(f"   ⚙️ LOCAL+SKU rule applied: LOCAL {local} + SKU {sku} → Proveedor {proveedor_forzado_norm} (FORCED)")
                    else:
                        # El proveedor forzado NO está en Full.xlsx para este SKU
                        # NO SE PUEDE CUMPLIR LA REGLA → No debe surtirse de ningún otro proveedor
                        row_error = row.copy()
                        row_error["OBSERVACION"] = (
                            str(row.get("CENTRO_COSTO", "")) + 
                            f"//REGLA ESPECIAL NO CUMPLIDA: Proveedor {proveedor_forzado_norm} no existe en Full.xlsx para SKU {sku}//" + 
                            str(row.get("NOMBRE_LUGAR", ""))
                        )
                        df_errors.append(row_error)
                        reglas_aplicadas_local += 1  # Contar como regla aplicada (aunque falló)
                        print(f"   ❌ LOCAL+SKU rule FAILED: LOCAL {local} + SKU {sku} → Proveedor {proveedor_forzado_norm} NOT in Full.xlsx")
                        continue
            
            # REGLA 2: Aplicar bloqueos por quiebre de stock
            # Solo si NO hay regla LOCAL+SKU forzada (si proveedor_forzado existe, ya se aplicó)
            if rules_manager and not proveedor_forzado and len(proveedores_disponibles) > 1:
                # Solo aplicar bloqueos si hay más de 1 proveedor
                proveedores_bloqueados = rules_manager.get_blocked_proveedores_for_sku(sku)
                
                if proveedores_bloqueados:
                    # Normalizar proveedores bloqueados
                    proveedores_bloqueados_norm = [p.replace('.0', '') if p.endswith('.0') else p for p in proveedores_bloqueados]
                    
                    # Filtrar proveedores bloqueados
                    proveedores_filtrados = [p for p in proveedores_disponibles if p not in proveedores_bloqueados_norm]
                    
                    if proveedores_filtrados:
                        # Hay proveedores alternativos, usar los no bloqueados
                        proveedores_disponibles = proveedores_filtrados
                        reglas_aplicadas_bloqueo += 1
                        print(f"   🚫 Stock block applied: SKU {sku} - blocked {proveedores_bloqueados_norm}, using {proveedores_filtrados}")
                    else:
                        # Todos los proveedores están bloqueados, mantener original
                        warnings.append(f"⚠️ All suppliers blocked for SKU {sku}, keeping all: {proveedores_disponibles}")
            
            elif rules_manager and not proveedor_forzado and len(proveedores_disponibles) == 1:
                # Solo 1 proveedor disponible y NO hay regla forzada - verificar si está bloqueado
                if rules_manager.is_blocked(sku, proveedores_disponibles[0]):
                    # Proveedor bloqueado y es el único → NO generar orden
                    row_error = row.copy()
                    row_error["OBSERVACION"] = (
                        str(row.get("CENTRO_COSTO", "")) + 
                        "//Bloqueado por Quiebre de Stock//" + 
                        str(row.get("NOMBRE_LUGAR", ""))
                    )
                    df_errors.append(row_error)
                    reglas_aplicadas_bloqueo += 1
                    print(f"   🚫 Order blocked: SKU {sku} + Proveedor {proveedores_disponibles[0]} (only supplier, blocked by stock rule)")
                    continue
            
            # Asignar proveedor (tomar el primero de los disponibles)
            proveedor_final = proveedores_disponibles[0]
            
            row_mapped = row.copy()
            row_mapped["PROVEEDOR"] = proveedor_final
            # Marcar si viene de regla especial para errores posteriores
            row_mapped["_REGLA_ESPECIAL"] = tiene_regla_especial
            df_mapped.append(row_mapped)
        
        df_mapped = pd.DataFrame(df_mapped) if df_mapped else pd.DataFrame(columns=df.columns.tolist() + ["PROVEEDOR"])
        df_errors = pd.DataFrame(df_errors) if df_errors else pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
        
        warnings.append(f"✅ Successfully mapped: {len(df_mapped)} records")
        warnings.append(f"⚠️ Sin precios/Bloqueados: {len(df_errors)} registros")
        
        # Resumen de reglas aplicadas
        if rules_manager and (reglas_aplicadas_local > 0 or reglas_aplicadas_bloqueo > 0):
            warnings.append(f"⚙️ Special rules applied:")
            if reglas_aplicadas_local > 0:
                warnings.append(f"   • LOCAL + SKU → Proveedor rules: {reglas_aplicadas_local} records")
            if reglas_aplicadas_bloqueo > 0:
                warnings.append(f"   • Stock block rules: {reglas_aplicadas_bloqueo} records")
        
    except Exception as e:
        warnings.append(f"❌ Error mapping suppliers: {e}")
        df_mapped = df.copy()
        df_mapped["PROVEEDOR"] = "ERROR"
        df_errors = pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
    
    return df_mapped, df_errors, warnings

# --- Procesamiento de fechas optimizado ---

def _fallback_fecha_processing(df, agenda_xlsm):
    """
    Método fallback para procesar fechas sin xlwings usando openpyxl
    """
    print("🔄 Using fallback method (openpyxl) for date processing...")
    
    try:
        from openpyxl import load_workbook
        
        # Intentar abrir con openpyxl (solo lectura)
        wb = load_workbook(agenda_xlsm, read_only=True, data_only=True)
        
        if "Matriz" not in wb.sheetnames:
            print("⚠️ Matriz sheet not found, using current date")
            fecha_formateada = datetime.now().strftime("%d/%m/%Y")
        else:
            ws = wb["Matriz"]
            fecha_m1 = ws["M1"].value
            
            if isinstance(fecha_m1, datetime):
                fecha_formateada = fecha_m1.strftime("%d/%m/%Y")
            else:
                try:
                    import pandas as pd
                    fecha_dt = pd.to_datetime(fecha_m1)
                    fecha_formateada = fecha_dt.strftime("%d/%m/%Y")
                except:
                    print("⚠️ Invalid date in M1, using current date")
                    fecha_formateada = datetime.now().strftime("%d/%m/%Y")
        
        wb.close()
        print(f"✅ Using delivery date: {fecha_formateada}")
        
        # Aplicar fecha y observaciones básicas
        df["FECHA_ENTREGA"] = fecha_formateada
        
        # Crear observaciones básicas por proveedor
        for proveedor in df["PROVEEDOR"].unique():
            mask_proveedor = df["PROVEEDOR"] == proveedor
            observacion_default = f"Delivery {fecha_formateada}"
            df.loc[mask_proveedor, "OBSERVACION"] = (
                df.loc[mask_proveedor, "CENTRO_COSTO"].fillna("").astype(str) + 
                "//" + observacion_default + "//" + 
                df.loc[mask_proveedor, "NOMBRE_LUGAR"].fillna("").astype(str)
            )
        
        print(f"✅ Fallback processing completed for {len(df)} records")
        return df, pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"] if "OBSERVACION" not in df.columns else df.columns)
        
    except Exception as e:
        print(f"❌ Fallback method also failed: {e}")
        # Último recurso: usar fecha actual
        fecha_fallback = datetime.now().strftime("%d/%m/%Y")
        df["FECHA_ENTREGA"] = fecha_fallback
        df["OBSERVACION"] = df.get("OBSERVACION", "") + f"//Error processing agenda, using {fecha_fallback}//"
        return df, pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"] if "OBSERVACION" not in df.columns else df.columns)

def rellenar_fecha_entrega_y_observacion_con_agenda_manager(df, fecha_pedido=None):
    """
    Rellena fecha de entrega y observación usando AgendaManager (sistema nuevo)
    
    Args:
        df: DataFrame con columnas PROVEEDOR, CENTRO_COSTO, NOMBRE_LUGAR
        fecha_pedido: Fecha del pedido (opcional, usa fecha actual si no se proporciona)
    
    Returns:
        Tuple (df_valid, df_err): DataFrames con registros válidos y con errores
    """
    try:
        from agenda_manager import AgendaManager
        
        print("📅 Processing dates with new AgendaManager system...")
        manager = AgendaManager()
        
        # Usar fecha actual si no se proporciona
        if fecha_pedido is None:
            fecha_pedido = datetime.now()
        elif isinstance(fecha_pedido, str):
            try:
                fecha_pedido = datetime.strptime(fecha_pedido, "%d-%m-%Y")
            except:
                try:
                    fecha_pedido = datetime.strptime(fecha_pedido, "%Y-%m-%d")
                except:
                    fecha_pedido = datetime.now()
        
        # Calcular fecha de despacho
        fecha_despacho = manager.calcular_fecha_despacho(fecha_pedido)
        dd_mm = fecha_despacho.strftime("%d-%m")
        
        print(f"📅 Order Date: {fecha_pedido.strftime('%d-%m-%Y')}")
        print(f"📅 Dispatch Date: {fecha_despacho.strftime('%d-%m-%Y')} (adding {manager.dias_despacho} days)")
        
        df = df.copy()
        df['FECHA_ENTREGA'] = None
        df['OBSERVACION'] = ''
        
        df_err = pd.DataFrame(columns=df.columns.tolist())
        df_valid_list = []
        df_err_list = []
        
        # Procesar cada registro
        proveedores_sin_config = set()
        
        for idx, row in df.iterrows():
            codigo_prov = str(row.get('PROVEEDOR', '')).strip()
            
            # Normalizar código (eliminar .0 si existe)
            codigo_prov = codigo_prov.replace('.0', '')
            
            if codigo_prov:
                fecha_entrega = manager.calcular_fecha_entrega(codigo_prov, fecha_despacho)
                
                if fecha_entrega:
                    # Proveedor configurado - registro válido
                    row_copy = row.copy()
                    row_copy['FECHA_ENTREGA'] = fecha_entrega.strftime("%d-%m-%Y")
                    
                    # Generar observación
                    centro_costo = str(row.get('CENTRO_COSTO', '')).strip()
                    nombre_lugar = limpiar_nombre_lugar(str(row.get('NOMBRE_LUGAR', '')))
                    row_copy['OBSERVACION'] = f"{centro_costo}//{dd_mm}//{nombre_lugar}"
                    
                    df_valid_list.append(row_copy)
                else:
                    # Proveedor no configurado - va a errores
                    proveedores_sin_config.add(codigo_prov)
                    row_copy = row.copy()
                    centro_costo = str(row.get('CENTRO_COSTO', '')).strip()
                    nombre_lugar = str(row.get('NOMBRE_LUGAR', '')).strip()
                    
                    # Verificar si viene de regla especial
                    tiene_regla = row.get('_REGLA_ESPECIAL', False)
                    if tiene_regla:
                        row_copy['OBSERVACION'] = f"{centro_costo}//REGLA ESPECIAL NO CUMPLIDA: Proveedor {codigo_prov} no está configurado en Agenda//{nombre_lugar}"
                    else:
                        row_copy['OBSERVACION'] = f"{centro_costo}//Falta Agenda//{nombre_lugar}"
                    df_err_list.append(row_copy)
            else:
                # Sin código de proveedor - va a errores
                row_copy = row.copy()
                row_copy['OBSERVACION'] = "//Sin código de proveedor//"
                df_err_list.append(row_copy)
        
        if proveedores_sin_config:
            print(f"⚠️ Suppliers not configured in agenda ({len(proveedores_sin_config)}):")
            for prov in sorted(proveedores_sin_config):
                print(f"   • {prov}")
        
        # Crear DataFrames de resultados con índices únicos
        if df_valid_list:
            df_valid = pd.DataFrame(df_valid_list).reset_index(drop=True)
            print(f"✅ {len(df_valid)} records with valid delivery dates")
        else:
            df_valid = pd.DataFrame(columns=df.columns.tolist() + ['FECHA_ENTREGA', 'OBSERVACION'])
        
        if df_err_list:
            df_err = pd.DataFrame(df_err_list).reset_index(drop=True)
            print(f"⚠️ {len(df_err)} records with errors (no agenda config)")
        else:
            df_err = pd.DataFrame(columns=df.columns.tolist() + ['OBSERVACION'])
        
        return df_valid, df_err
        
    except Exception as e:
        print(f"❌ Error using AgendaManager: {e}")
        # Retornar todos como errores
        df_err = df.copy()
        df_err['OBSERVACION'] = "//Error en sistema de agenda//"
        return pd.DataFrame(columns=df.columns.tolist() + ['FECHA_ENTREGA', 'OBSERVACION']), df_err


def rellenar_fecha_entrega_y_observacion(df, agenda_xlsm=None):
    """
    Rellena fecha de entrega y observación usando el nuevo sistema AgendaManager
    
    PROCESO:
    - Busca cada proveedor en la configuración de agenda (agenda_config.json)
    - Calcula fecha de entrega automáticamente según días configurados
    - Si hay fecha manual, la usa en lugar del cálculo
    - Observación con fecha de despacho en formato dd-mm + BOD. + lugar
    
    Args:
        df: DataFrame con columnas PROVEEDOR, CENTRO_COSTO, NOMBRE_LUGAR
        agenda_xlsm: Parámetro legacy para compatibilidad (ya no se usa)
    
    Returns:
        Tuple (df_valid, df_err): DataFrames con registros válidos y con errores
    
    NOTA: El parámetro agenda_xlsm se mantiene para compatibilidad pero ya no se usa.
          Todo el procesamiento se hace con AgendaManager (agenda_config.json)
    """
    print("📅 Processing dates with AgendaManager system (Python-based)...")
    
    # Usar directamente el nuevo sistema
    try:
        df_valid, df_err = rellenar_fecha_entrega_y_observacion_con_agenda_manager(df)
        return df_valid, df_err
    except Exception as e:
        print(f"❌ Error in AgendaManager: {e}")
        # Si falla, retornar todo como errores
        df_err = df.copy()
        df_err['OBSERVACION'] = "//Error en sistema de agenda//"
        return pd.DataFrame(columns=df.columns.tolist() + ['FECHA_ENTREGA', 'OBSERVACION']), df_err
    
    df = df.copy()
    df_err = pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"] if "OBSERVACION" not in df.columns else df.columns)
    
    app = None
    wb = None
    
    try:
        # Verificar que el archivo exista
        if not os.path.exists(agenda_xlsm):
            print(f"❌ Agenda.xlsm not found: {agenda_xlsm}")
            # Todos van a errores si no hay archivo
            df_err_agenda = df.copy()
            df_err_agenda["OBSERVACION"] = df_err_agenda["CENTRO_COSTO"].fillna("") + "//Falta Agenda//" + df_err_agenda["NOMBRE_LUGAR"].fillna("")
            return pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"]), df_err_agenda
        
        print("🔗 Opening Agenda.xlsm with xlwings...")
        
        try:
            app = xw.App(visible=False, add_book=False)
            wb = app.books.open(agenda_xlsm)
            
        except Exception as e:
            print(f"❌ Error opening with xlwings: {e}")
            print("🔄 Trying openpyxl fallback...")
            return _procesar_agenda_con_openpyxl_correcto(df, agenda_xlsm)
        
        # Verificar si existe la hoja "Matriz"
        sheet_names = [sheet.name for sheet in wb.sheets]
        if "Matriz" not in sheet_names:
            print("❌ Sheet 'Matriz' not found in Agenda.xlsm")
            df_err_agenda = df.copy()
            df_err_agenda["OBSERVACION"] = df_err_agenda["CENTRO_COSTO"].fillna("") + "//Falta Agenda//" + df_err_agenda["NOMBRE_LUGAR"].fillna("")
            return pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"]), df_err_agenda
        
        ws_matriz = wb.sheets["Matriz"]
        
        # Obtener fecha de despacho de M2 (para observaciones)
        print("📅 Reading dispatch date from M2...")
        fecha_despacho = ws_matriz.range("M2").value
        print(f"📅 Raw dispatch date from M2: {fecha_despacho}")
        
        if not isinstance(fecha_despacho, datetime):
            try:
                fecha_despacho = pd.to_datetime(fecha_despacho)
            except:
                fecha_despacho = datetime.today()
                print("⚠️ Using current date as dispatch date")
        
        dd_mm = fecha_despacho.strftime("%d-%m")
        print(f"✅ Using dispatch date for observations: {dd_mm}")
        
        # Leer matriz de proveedores y fechas de entrega
        print("📋 Reading supplier delivery matrix...")
        ultima_fila = ws_matriz.range("A" + str(ws_matriz.cells.last_cell.row)).end("up").row
        print(f"📊 Matrix data range: A3:K{ultima_fila}")
        
        if ultima_fila >= 3:
            data_matriz = ws_matriz.range(f"A3:K{ultima_fila}").value
            columnas = ["PROVEEDOR", "NOMBRE", "VACIA", "LUN", "MAR", "MIE", "JUE", "VIE", "SAB", "D-1", "ENTREGA"]
            
            # Convertir a DataFrame
            if isinstance(data_matriz[0], (list, tuple)):
                df_matriz = pd.DataFrame(data_matriz, columns=columnas)
            else:
                # Solo una fila
                df_matriz = pd.DataFrame([data_matriz], columns=columnas)
        else:
            print("⚠️ No data found in matrix")
            df_matriz = pd.DataFrame(columns=["PROVEEDOR", "ENTREGA"])
        
        # Limpiar y procesar datos de la matriz
        df_matriz["PROVEEDOR"] = df_matriz["PROVEEDOR"].astype(str).str.strip()
        # Normalizar códigos de proveedor eliminando .0 si existe
        df_matriz["PROVEEDOR"] = df_matriz["PROVEEDOR"].str.replace(r'\.0$', '', regex=True)
        df_matriz = df_matriz[df_matriz["PROVEEDOR"] != "nan"]  # Eliminar filas vacías
        df_matriz = df_matriz[df_matriz["PROVEEDOR"] != ""]
        
        print(f"📋 Found {len(df_matriz)} suppliers in matrix:")
        for idx, row in df_matriz.iterrows():
            print(f"   • {row['PROVEEDOR']}: {row['ENTREGA']}")
        
        # Procesar fechas de entrega
        try:
            df_matriz["ENTREGA"] = pd.to_datetime(df_matriz["ENTREGA"], format="%d-%m-%Y", errors="coerce")
            df_matriz["FECHA_ENTREGA"] = df_matriz["ENTREGA"].dt.strftime("%d-%m-%Y")
        except Exception as e:
            print(f"⚠️ Error processing delivery dates: {e}")
            df_matriz["FECHA_ENTREGA"] = datetime.today().strftime("%d-%m-%Y")
        
        # Hacer merge con los datos
        print("🔗 Merging suppliers with delivery dates...")
        print(f"📊 Input data has {len(df)} records with {len(df['PROVEEDOR'].unique())} unique suppliers:")
        for prov in sorted(df["PROVEEDOR"].unique()):
            count = len(df[df["PROVEEDOR"] == prov])
            print(f"   • '{prov}': {count} records")
        
        print(f"📋 Agenda has {len(df_matriz)} suppliers:")
        for idx, row in df_matriz.iterrows():
            print(f"   • '{row['PROVEEDOR']}': {row['FECHA_ENTREGA']}")
        
        df_merged = df.merge(df_matriz[["PROVEEDOR", "FECHA_ENTREGA"]], on="PROVEEDOR", how="left")
        
        # Separar válidos de errores
        mask_falta_agenda = df_merged["FECHA_ENTREGA"].isna()
        
        if mask_falta_agenda.any():
            df_err_agenda = df_merged.loc[mask_falta_agenda].copy()
            df_err_agenda["OBSERVACION"] = (
                df_err_agenda["CENTRO_COSTO"].fillna("") + 
                "//Falta Agenda//" + 
                df_err_agenda["NOMBRE_LUGAR"].fillna("")
            )
            df_err = pd.concat([df_err, df_err_agenda], ignore_index=True)
            
            proveedores_sin_agenda = df_err_agenda["PROVEEDOR"].unique()
            print(f"⚠️ Suppliers not found in agenda ({len(proveedores_sin_agenda)}):")
            for prov in proveedores_sin_agenda:
                print(f"   • {prov}")
        
        # Procesar registros válidos
        df_valid = df_merged.loc[~mask_falta_agenda].copy()
        
        if len(df_valid) > 0:
            # Limpiar NOMBRE_LUGAR usando la función auxiliar
            df_valid["NOMBRE_LUGAR_LIMPIO"] = df_valid["NOMBRE_LUGAR"].apply(limpiar_nombre_lugar)
            
            # Crear observaciones en el formato correcto: CENTRO_COSTO//dd-mm//NOMBRE_LUGAR_LIMPIO
            df_valid["OBSERVACION"] = (
                df_valid["CENTRO_COSTO"].fillna("") + 
                f"//{dd_mm}//" + 
                df_valid["NOMBRE_LUGAR_LIMPIO"]
            )
            
            print(f"✅ Successfully processed {len(df_valid)} records with delivery dates")
            print(f"📊 Sample observations:")
            for idx, row in df_valid.head(3).iterrows():
                print(f"   • {row['PROVEEDOR']}: {row['FECHA_ENTREGA']} | {row['OBSERVACION']}")
        else:
            print("⚠️ No valid records found")
        
        print(f"📊 Summary: {len(df_valid)} valid, {len(df_err)} errors")
        
    except Exception as e:
        print(f"❌ Error processing Agenda.xlsm: {e}")
        # Todos van a errores en caso de error crítico
        df_err_agenda = df.copy()
        df_err_agenda["OBSERVACION"] = df_err_agenda["CENTRO_COSTO"].fillna("") + "//Error Agenda//" + df_err_agenda["NOMBRE_LUGAR"].fillna("")
        return pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"]), df_err_agenda
        
    finally:
        # Cerrar Excel de manera segura
        try:
            if wb:
                wb.close()
            if app:
                app.quit()
        except Exception as e:
            print(f"⚠️ Warning closing Excel: {e}")
    
    return df_valid.reset_index(drop=True), df_err.reset_index(drop=True)


def _procesar_agenda_con_openpyxl_correcto(df, agenda_xlsm):
    """Fallback correcto: procesar agenda con openpyxl siguiendo el proceso original"""
    print("� Using openpyxl fallback method...")
    
    try:
        wb = load_workbook(agenda_xlsm, read_only=True)
        
        if 'Matriz' not in wb.sheetnames:
            print("❌ Matriz sheet not found in openpyxl fallback")
            df_err = df.copy()
            # Limpiar también aquí
            df_err["OBSERVACION"] = df_err["CENTRO_COSTO"].fillna("") + "//Falta Agenda//" + df_err["NOMBRE_LUGAR"].apply(limpiar_nombre_lugar)
            return pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"]), df_err
        
        ws = wb['Matriz']
        
        # Obtener fecha de despacho de M2
        fecha_despacho = ws['M2'].value
        if not isinstance(fecha_despacho, datetime):
            try:
                fecha_despacho = pd.to_datetime(fecha_despacho)
            except:
                fecha_despacho = datetime.today()
        
        dd_mm = fecha_despacho.strftime("%d-%m")
        print(f"✅ Fallback - dispatch date: {dd_mm}")
        
        # Leer matriz de proveedores (A3 hacia abajo)
        proveedores_agenda = {}
        
        # Buscar hasta fila 100 (ajustar según necesidad)
        for row in range(3, 101):
            try:
                proveedor = ws[f'A{row}'].value
                entrega = ws[f'K{row}'].value  # Columna K = ENTREGA
                
                if proveedor and str(proveedor).strip() and str(proveedor).strip() != 'None':
                    proveedor_clean = str(proveedor).strip()
                    
                    if entrega and isinstance(entrega, datetime):
                        fecha_entrega = entrega.strftime("%d-%m-%Y")
                        proveedores_agenda[proveedor_clean] = fecha_entrega
                        print(f"   • {proveedor_clean}: {fecha_entrega}")
                    elif entrega and str(entrega).strip():
                        try:
                            fecha_dt = pd.to_datetime(str(entrega), format="%d-%m-%Y")
                            fecha_entrega = fecha_dt.strftime("%d-%m-%Y")
                            proveedores_agenda[proveedor_clean] = fecha_entrega
                            print(f"   • {proveedor_clean}: {fecha_entrega}")
                        except:
                            print(f"   ⚠️ {proveedor_clean}: invalid date format")
                            
            except Exception as e:
                continue  # Fila vacía o error, continuar
        
        wb.close()
        
        print(f"📋 Fallback found {len(proveedores_agenda)} suppliers in agenda")
        
        # Aplicar fechas de entrega
        df_valid = []
        df_errors = []
        
        for idx, row in df.iterrows():
            proveedor = str(row["PROVEEDOR"]).strip()
            
            if proveedor in proveedores_agenda:
                # Proveedor encontrado - válido
                row_copy = row.copy()
                row_copy["FECHA_ENTREGA"] = proveedores_agenda[proveedor]
                row_copy["OBSERVACION"] = (
                    str(row["CENTRO_COSTO"]) + 
                    f"//{dd_mm}//" + 
                    str(row["NOMBRE_LUGAR"])
                )
                df_valid.append(row_copy)
            else:
                # Proveedor no encontrado - error
                row_copy = row.copy()
                row_copy["OBSERVACION"] = (
                    str(row["CENTRO_COSTO"]) + 
                    "//Falta Agenda//" + 
                    str(row["NOMBRE_LUGAR"])
                )
                df_errors.append(row_copy)
        
        df_valid = pd.DataFrame(df_valid) if df_valid else pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"])
        df_errors = pd.DataFrame(df_errors) if df_errors else pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
        
        print(f"✅ Fallback processing: {len(df_valid)} valid, {len(df_errors)} errors")
        
        return df_valid, df_errors
        
    except Exception as e:
        print(f"❌ Fallback method failed: {e}")
        # Último recurso - todos van a errores
        df_err = df.copy()
        df_err["OBSERVACION"] = df_err["CENTRO_COSTO"].fillna("") + "//Error Agenda//" + df_err["NOMBRE_LUGAR"].fillna("")
        return pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"]), df_err


def _usar_fecha_fallback(df, error_msg):
    """Usar fecha actual como fallback cuando fallan todos los métodos"""
    print(f"📅 Using fallback date due to: {error_msg}")
    fecha_actual = datetime.now().strftime("%d/%m/%Y")
    df["FECHA_ENTREGA"] = fecha_actual
    df["OBSERVACION"] = df.get("OBSERVACION", "").fillna("").astype(str) + f"//Fallback date {fecha_actual}//"
    df_err = pd.DataFrame(columns=df.columns.tolist())
    return df, df_err

# --- Generación de nombre de archivo optimizada ---

def obtener_nombre_archivo_salida(agenda_xlsm, base_dir):
    """
    Genera el nombre del archivo de salida basado en la fecha de M1 en Agenda.xlsm
    Versión optimizada con mejor manejo de errores y fallback
    """
    print(f"📅 Getting output filename from: {agenda_xlsm}")
    
    app = None
    wb = None
    
    try:
        # Verificar que el archivo exista
        if not os.path.exists(agenda_xlsm):
            print(f"⚠️ Agenda.xlsm not found, using current date")
            fecha_actual = datetime.now().strftime("%d-%m-%Y")
            return _crear_ruta_salida(base_dir, fecha_actual)
        
        # Intentar con xlwings primero (con timeout)
        try:
            import time
            start_time = time.time()
            timeout = 15  # 15 segundos timeout para esta función
            
            # Limpiar apps existentes
            try:
                for existing_app in xw.apps:
                    existing_app.quit()
            except:
                pass
            
            app = xw.App(visible=False, add_book=False)
            app.display_alerts = False
            
            if time.time() - start_time > timeout:
                raise TimeoutError("Excel opening timeout")
            
            wb = app.books.open(agenda_xlsm)
            
            if "Matriz" not in [sheet.name for sheet in wb.sheets]:
                print(f"⚠️ Matriz sheet not found, using current date")
                fecha_actual = datetime.now().strftime("%d-%m-%Y")
                return _crear_ruta_salida(base_dir, fecha_actual)
            
            ws_matriz = wb.sheets["Matriz"]
            fecha_m2 = ws_matriz.range("M2").value
            
            # Procesar fecha
            if isinstance(fecha_m2, datetime):
                fecha_str = fecha_m2.strftime("%d-%m-%Y")
            else:
                try: 
                    fecha_dt = pd.to_datetime(fecha_m2)
                    fecha_str = fecha_dt.strftime("%d-%m-%Y")
                except: 
                    print(f"⚠️ Invalid date in M2: {fecha_m2}, using current date")
                    fecha_str = datetime.now().strftime("%d-%m-%Y")
            
            print(f"✅ Using date from M2: {fecha_str}")
            return _crear_ruta_salida(base_dir, fecha_str)
            
        except Exception as xlwings_error:
            print(f"⚠️ xlwings failed: {xlwings_error}")
            print("🔄 Trying fallback with openpyxl...")
            
            # Fallback con openpyxl
            try:
                from openpyxl import load_workbook
                wb_openpyxl = load_workbook(agenda_xlsm, read_only=True, data_only=True)
                
                if "Matriz" not in wb_openpyxl.sheetnames:
                    fecha_str = datetime.now().strftime("%d-%m-%Y")
                else:
                    ws = wb_openpyxl["Matriz"]
                    fecha_m2 = ws["M2"].value
                    
                    if isinstance(fecha_m2, datetime):
                        fecha_str = fecha_m2.strftime("%d-%m-%Y")
                    else:
                        try:
                            fecha_dt = pd.to_datetime(fecha_m2)
                            fecha_str = fecha_dt.strftime("%d-%m-%Y")
                        except:
                            fecha_str = datetime.now().strftime("%d-%m-%Y")
                
                wb_openpyxl.close()
                print(f"✅ Fallback successful, using date: {fecha_str}")
                return _crear_ruta_salida(base_dir, fecha_str)
                
            except Exception as openpyxl_error:
                print(f"⚠️ Openpyxl fallback also failed: {openpyxl_error}")
                fecha_actual = datetime.now().strftime("%d-%m-%Y")
                return _crear_ruta_salida(base_dir, fecha_actual)
        
    except Exception as e:
        print(f"⚠️ General error: {e}, using current date")
        fecha_actual = datetime.now().strftime("%d-%m-%Y")
        return _crear_ruta_salida(base_dir, fecha_actual)
        
    finally:
        # Cerrar Excel de manera segura
        try:
            if wb:
                wb.close()
            if app:
                app.quit()
        except Exception as e:
            print(f"⚠️ Warning closing Excel: {e}")
            # Forzar cierre si es necesario
            try:
                import subprocess
                subprocess.run(["taskkill", "/f", "/im", "EXCEL.EXE"], capture_output=True)
            except:
                pass

def _crear_ruta_salida(base_dir, fecha_str):
    """Crea la ruta completa del archivo de salida"""
    nombre_archivo = f"PEDIDOS_CD_OVIEDO_{fecha_str}.xlsx"
    salidas_dir = os.path.join(base_dir, "Salidas")
    os.makedirs(salidas_dir, exist_ok=True)
    return os.path.join(salidas_dir, nombre_archivo)

# --- Asignación de IDs optimizada ---

def asignar_id_final(df):
    """
    Asigna IDs finales consolidando duplicados
    Versión optimizada con mejor logging y validación
    """
    print("🏷️ Assigning final IDs and consolidating duplicates...")
    
    df = df.copy()
    
    # NO eliminar _REGLA_ESPECIAL aquí - se necesita para errores de agenda
    # Se eliminará justo antes de guardar el Excel
    
    # Validar columnas requeridas
    columnas_requeridas = ["LOCAL", "SKU", "PROVEEDOR", "FECHA_ENTREGA", "OBSERVACION", "CANTIDAD"]
    columnas_faltantes = [col for col in columnas_requeridas if col not in df.columns]
    
    if columnas_faltantes:
        print(f"⚠️ Missing columns: {columnas_faltantes}")
        for col in columnas_faltantes:
            df[col] = ""
    
    print(f"📊 Input records: {len(df)}")
    
    # Consolidar duplicados
    columnas_agrupacion = ["LOCAL", "SKU", "PROVEEDOR", "FECHA_ENTREGA", "OBSERVACION"]
    
    try:
        # Agrupar y sumar cantidades
        df_consolidado = df.groupby(columnas_agrupacion, as_index=False).agg({
            'CANTIDAD': 'sum',
            'CENTRO_COSTO': 'first',
            'NOMBRE_LUGAR': 'first',
            '_SRC_FILE': lambda x: ', '.join(x.unique()) if len(x.unique()) > 1 else x.iloc[0]
        })
        
        print(f"✅ Consolidated to {len(df_consolidado)} unique records")
        df = df_consolidado
        
    except Exception as e:
        print(f"⚠️ Error consolidating duplicates: {e}")
        print("📝 Continuing without consolidation")
    
    # Asignar IDs por proveedor y observación
    df = df.sort_values(["PROVEEDOR", "OBSERVACION", "SKU"]).reset_index(drop=True)
    
    ids = []
    next_id = 1
    last_pair = (None, None)
    
    for _, row in df.iterrows():
        current_pair = (row["PROVEEDOR"], row["OBSERVACION"])
        if current_pair != last_pair:
            ids.append(next_id)
            last_pair = current_pair
            next_id += 1
        else:
            ids.append(next_id - 1)
    
    df["ID PEDIDO"] = ids
    
    # Verificar distribución de IDs
    id_counts = df["ID PEDIDO"].value_counts().sort_index()
    print(f"📋 ID distribution: {len(id_counts)} unique order IDs")
    print(f"📊 Records per ID - Min: {id_counts.min()}, Max: {id_counts.max()}, Avg: {id_counts.mean():.1f}")
    
    # Reordenar columnas
    columnas_finales = ["ID PEDIDO", "LOCAL", "PROVEEDOR", "FECHA_ENTREGA", "SKU", "CANTIDAD", "OBSERVACION"]
    
    # Asegurar que todas las columnas finales existan
    for col in columnas_finales:
        if col not in df.columns:
            df[col] = ""
    
    df_final = df[columnas_finales].copy()
    
    print(f"✅ Final output: {len(df_final)} records with {df_final['ID PEDIDO'].nunique()} order IDs")
    
    return df_final

# --- Formateo de Excel optimizado ---

def formatear_excel_salida(archivo_excel):
    """
    Formatea el archivo Excel con estilos profesionales DHL
    Versión optimizada con mejor diseño y manejo de errores
    """
    print(f"🎨 Applying professional formatting to: {os.path.basename(archivo_excel)}")
    
    try:
        wb = load_workbook(archivo_excel)
        
        # Formatear hoja principal "PEDIDOS_CD"
        if "PEDIDOS_CD" in wb.sheetnames:
            _formatear_hoja_pedidos(wb["PEDIDOS_CD"])
            print("✅ PEDIDOS_CD sheet formatted")
        
        # Formatear hoja de errores si existe
        if "Errores" in wb.sheetnames:
            _formatear_hoja_errores(wb["Errores"])
            print("✅ Errores sheet formatted")
        elif "Errors" in wb.sheetnames:
            _formatear_hoja_errores(wb["Errors"])
            print("✅ Errors sheet formatted")
        
        # Guardar cambios
        wb.save(archivo_excel)
        print(f"✅ Formatted file saved: {os.path.basename(archivo_excel)}")
        
    except Exception as e:
        print(f"⚠️ Error formatting Excel: {e}")
        print("📄 File saved without special formatting")

def _formatear_hoja_pedidos(ws):
    """Formatear hoja de pedidos con estilo DHL"""
    # Definir colores DHL
    dhl_red = "D40511"
    dhl_yellow = "FFCC00"
    dhl_dark = "1A1A1A"
    
    # Anchos de columna optimizados
    anchos_columnas = {
        'A': 12,  # ID PEDIDO
        'B': 8,   # LOCAL
        'C': 15,  # PROVEEDOR
        'D': 15,  # FECHA_ENTREGA
        'E': 12,  # SKU
        'F': 10,  # CANTIDAD
        'G': 70   # OBSERVACION
    }
    
    # Aplicar anchos
    for col, ancho in anchos_columnas.items():
        ws.column_dimensions[col].width = ancho
    
    # Estilo para encabezados
    header_font = Font(bold=True, color="FFFFFF", size=11)
    header_fill = PatternFill(start_color=dhl_red, end_color=dhl_red, fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Aplicar formato a encabezados
    for cell in ws[1]:
        if cell.value:
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
    
    # Estilo para datos
    data_font = Font(size=10)
    data_alignment_center = Alignment(horizontal="center", vertical="center")
    data_alignment_left = Alignment(horizontal="left", vertical="center", wrap_text=True)
    
    # Alternar colores de filas
    light_fill = PatternFill(start_color="F8F9FA", end_color="F8F9FA", fill_type="solid")
    
    # Aplicar formato a datos
    for row_num, row in enumerate(ws.iter_rows(min_row=2), start=2):
        for col_num, cell in enumerate(row):
            cell.font = data_font
            
            # Alineación por columna
            if col_num in [0, 1, 2, 3, 4, 5]:  # ID, LOCAL, PROVEEDOR, FECHA, SKU, CANTIDAD
                cell.alignment = data_alignment_center
            elif col_num == 6:  # OBSERVACION
                cell.alignment = data_alignment_left
            
            # Alternar color de filas
            if row_num % 2 == 0:
                cell.fill = light_fill

def _formatear_hoja_errores(ws):
    """Formatear hoja de errores"""
    # Anchos para hoja de errores
    anchos_errores = {
        'A': 8,   # LOCAL
        'B': 12,  # SKU
        'C': 10,  # CANTIDAD
        'D': 15,  # CENTRO_COSTO
        'E': 30,  # NOMBRE_LUGAR
        'F': 25,  # _SRC_FILE
        'G': 60   # OBSERVACION
    }
    
    # Aplicar anchos
    for col, ancho in anchos_errores.items():
        ws.column_dimensions[col].width = ancho
    
    # Estilo para encabezados de errores
    header_font_error = Font(bold=True, color="FFFFFF", size=11)
    header_fill_error = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Formatear encabezados
    for cell in ws[1]:
        if cell.value:
            cell.font = header_font_error
            cell.fill = header_fill_error
            cell.alignment = header_alignment
    
    # Formatear datos
    data_font = Font(size=9)
    data_alignment = Alignment(horizontal="left", vertical="center", wrap_text=True)
    error_fill = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
    
    for row in ws.iter_rows(min_row=2):
        for cell in row:
            cell.font = data_font
            cell.alignment = data_alignment
            cell.fill = error_fill

# --- Funciones fallback para problemas con xlwings ---

def _procesar_agenda_con_openpyxl_correcto(df, agenda_xlsm):
    """Fallback correcto: procesar agenda con openpyxl siguiendo el proceso original"""
    print("📚 Using openpyxl fallback method...")
    
    try:
        wb = load_workbook(agenda_xlsm, read_only=True)
        
        if 'Matriz' not in wb.sheetnames:
            print("❌ Matriz sheet not found in openpyxl fallback")
            df_err = df.copy()
            df_err["OBSERVACION"] = df_err["CENTRO_COSTO"].fillna("") + "//Falta Agenda//" + df_err["NOMBRE_LUGAR"].fillna("")
            return pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"]), df_err
        
        ws = wb['Matriz']
        
        # Obtener fecha de despacho de M2
        fecha_despacho = ws['M2'].value
        if not isinstance(fecha_despacho, datetime):
            try:
                fecha_despacho = pd.to_datetime(fecha_despacho)
            except:
                fecha_despacho = datetime.today()
        
        dd_mm = fecha_despacho.strftime("%d-%m")
        print(f"✅ Fallback - dispatch date: {dd_mm}")
        
        # Leer matriz de proveedores (A3 hacia abajo)
        proveedores_agenda = {}
        
        # Buscar hasta fila 100 (ajustar según necesidad)
        for row in range(3, 101):
            try:
                proveedor = ws[f'A{row}'].value
                entrega = ws[f'K{row}'].value  # Columna K = ENTREGA
                
                if proveedor and str(proveedor).strip() and str(proveedor).strip() != 'None':
                    proveedor_clean = str(proveedor).strip()
                    
                    if entrega and isinstance(entrega, datetime):
                        fecha_entrega = entrega.strftime("%d-%m-%Y")
                        proveedores_agenda[proveedor_clean] = fecha_entrega
                        print(f"   • {proveedor_clean}: {fecha_entrega}")
                    elif entrega and str(entrega).strip():
                        try:
                            fecha_dt = pd.to_datetime(str(entrega), format="%d-%m-%Y")
                            fecha_entrega = fecha_dt.strftime("%d-%m-%Y")
                            proveedores_agenda[proveedor_clean] = fecha_entrega
                            print(f"   • {proveedor_clean}: {fecha_entrega}")
                        except:
                            print(f"   ⚠️ {proveedor_clean}: invalid date format")
                            
            except Exception as e:
                continue  # Fila vacía o error, continuar
        
        wb.close()
        
        print(f"📋 Fallback found {len(proveedores_agenda)} suppliers in agenda")
        
        # Aplicar fechas de entrega
        df_valid = []
        df_errors = []
        
        for idx, row in df.iterrows():
            proveedor = str(row["PROVEEDOR"]).strip()
            
            if proveedor in proveedores_agenda:
                # Proveedor encontrado - válido
                row_copy = row.copy()
                row_copy["FECHA_ENTREGA"] = proveedores_agenda[proveedor]
                row_copy["OBSERVACION"] = (
                    str(row["CENTRO_COSTO"]) + 
                    f"//{dd_mm}//" + 
                    str(row["NOMBRE_LUGAR"])
                )
                df_valid.append(row_copy)
            else:
                # Proveedor no encontrado - error
                row_copy = row.copy()
                row_copy["OBSERVACION"] = (
                    str(row["CENTRO_COSTO"]) + 
                    "//Falta Agenda//" + 
                    str(row["NOMBRE_LUGAR"])
                )
                df_errors.append(row_copy)
        
        df_valid = pd.DataFrame(df_valid) if df_valid else pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"])
        df_errors = pd.DataFrame(df_errors) if df_errors else pd.DataFrame(columns=df.columns.tolist() + ["OBSERVACION"])
        
        print(f"✅ Fallback processing: {len(df_valid)} valid, {len(df_errors)} errors")
        
        return df_valid, df_errors
        
    except Exception as e:
        print(f"❌ Fallback method failed: {e}")
        # Último recurso - todos van a errores
        df_err = df.copy()
        df_err["OBSERVACION"] = df_err["CENTRO_COSTO"].fillna("") + "//Error Agenda//" + df_err["NOMBRE_LUGAR"].fillna("")
        return pd.DataFrame(columns=df.columns.tolist() + ["FECHA_ENTREGA", "OBSERVACION"]), df_err

def _usar_fecha_fallback(df, reason):
    """Usar fecha actual como último recurso"""
    fecha_fallback = datetime.now().strftime("%d/%m/%Y")
    print(f"🔄 Using fallback date: {fecha_fallback}")
    print(f"📝 Reason: {reason}")
    
    df["FECHA_ENTREGA"] = fecha_fallback
    
    observacion_fallback = f"Delivery {fecha_fallback}"
    proveedores_unicos = df["PROVEEDOR"].unique()
    
    for proveedor in proveedores_unicos:
        mask_proveedor = df["PROVEEDOR"] == proveedor
        df.loc[mask_proveedor, "OBSERVACION"] = (
            df.loc[mask_proveedor, "CENTRO_COSTO"].fillna("").astype(str) + 
            "//" + observacion_fallback + "//" + 
            df.loc[mask_proveedor, "NOMBRE_LUGAR"].fillna("").astype(str)
        )
    
    print(f"✅ Fallback processing completed for {len(df)} records")
    
    df_err = pd.DataFrame(columns=df.columns.tolist())
    return df, df_err

# --- Función de utilidad para logging ---

def log_processing_step(step_name, status="INFO", details=""):
    """Función auxiliar para logging consistente"""
    timestamp = datetime.now().strftime("%H:%M:%S")
    icons = {
        "INFO": "ℹ️",
        "SUCCESS": "✅", 
        "WARNING": "⚠️",
        "ERROR": "❌",
        "START": "🚀",
        "PROCESS": "🔄"
    }
    
    icon = icons.get(status, "📝")
    message = f"[{timestamp}] {icon} {step_name}"
    if details:
        message += f": {details}"
    
    print(message)

def debug_supplier_matching(df, proveedores_agenda):
    """Función para debuggear el matching de proveedores"""
    proveedores_input = set(df["PROVEEDOR"].str.strip().unique())
    proveedores_agenda_set = set(proveedores_agenda.keys())
    
    print("🔍 SUPPLIER MATCHING DEBUG:")
    print(f"📊 Input suppliers ({len(proveedores_input)}):")
    for prov in sorted(proveedores_input):
        count = len(df[df["PROVEEDOR"].str.strip() == prov])
        print(f"   • '{prov}': {count} records")
    
    print(f"📋 Agenda suppliers ({len(proveedores_agenda_set)}):")
    for prov in sorted(proveedores_agenda_set):
        print(f"   • '{prov}': {proveedores_agenda[prov]}")
    
    # Mostrar coincidencias y diferencias
    encontrados = proveedores_input.intersection(proveedores_agenda_set)
    no_encontrados = proveedores_input - proveedores_agenda_set
    
    print(f"✅ MATCHES ({len(encontrados)}):")
    for prov in sorted(encontrados):
        print(f"   • '{prov}'")
    
    print(f"❌ NOT FOUND IN AGENDA ({len(no_encontrados)}):")
    for prov in sorted(no_encontrados):
        count = len(df[df["PROVEEDOR"].str.strip() == prov])
        print(f"   • '{prov}': {count} records")

if __name__ == "__main__":
    print("🚚 DHL Order Processing System - Processing Module v2.0")
    print("💻 Created by Lucas Gnemmi")
    print("📝 This module contains optimized functions for order processing")


def ajustar_cantidades_formato_minimo(df):
    """
    Ajusta las cantidades según el formato de empaque definido en cada SKU
    Calcula múltiplos del formato (ej: formato 60, pido 100 = 2*60 = 120)
    
    Args:
        df: DataFrame con columnas SKU y CANTIDAD
        
    Returns:
        DataFrame con cantidades ajustadas y registro de cambios
    """
    print("🔧 Applying packaging format adjustments...")
    
    # Verificar que existan las columnas necesarias
    if 'SKU' not in df.columns or 'CANTIDAD' not in df.columns:
        print("⚠️ SKU or CANTIDAD columns not found, skipping format adjustment")
        return df
    
    # Importar ProductsManager
    try:
        from products_manager import ProductsManager
    except ImportError:
        print("⚠️ ProductsManager not available, skipping format adjustment")
        return df
    
    # Inicializar ProductsManager
    products_manager = ProductsManager()
    df_adjusted = df.copy()
    adjustments_count = 0
    
    import math
    
    for index, row in df_adjusted.iterrows():
        sku = str(row['SKU']).strip().upper()
        cantidad_original = row['CANTIDAD']
        
        try:
            # Verificar si el SKU tiene formato de empaque
            formato_empaque = products_manager.get_formato_minimo(sku)
            
            if formato_empaque is not None and formato_empaque > 0:
                # Convertir cantidad a float para cálculo
                cantidad_float = float(cantidad_original)
                
                # Calcular formatos necesarios y cantidad final
                formatos_necesarios = math.ceil(cantidad_float / formato_empaque)
                cantidad_ajustada = formatos_necesarios * formato_empaque
                
                # Aplicar ajuste si es diferente
                if cantidad_ajustada != cantidad_float:
                    df_adjusted.at[index, 'CANTIDAD'] = cantidad_ajustada
                    adjustments_count += 1
                    print(f"   📦 SKU {sku}: {cantidad_original} → {cantidad_ajustada} ({formatos_necesarios} x {formato_empaque})")
                    
        except (ValueError, TypeError) as e:
            print(f"   ⚠️ Error processing SKU {sku}: {e}")
            continue
    
    if adjustments_count > 0:
        print(f"✅ Applied {adjustments_count} packaging format adjustments")
    else:
        print("✅ No format adjustments needed")
    
    return df_adjusted
